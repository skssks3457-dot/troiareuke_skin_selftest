from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
SUMMARY_SHEET = "\uC6D4\uAC04\uC694\uC57D"
MONTH_SUFFIX = "\uC6D4"
RAW_PREFIX = "\uC6D0\uBCF8_"
TOTAL_LABEL = "\uC804\uCCB4 \uC18C\uACC4"
TA_TOTAL_LABEL = "\uD2B8\uB85C\uC774\uC544\uB974\uCF00 \uC804\uCCB4"
SUBTOTAL_SUFFIX = " \uC18C\uACC4"
EXCLUDED_TERMS = ("\uD560\uC778", "\uD0DD\uBC30\uBE44")

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
HEADER_FILL = PatternFill("solid", fgColor="EAF1FB")
SECTION_FILL = PatternFill("solid", fgColor="F7F9FC")


def latest_program_file() -> Path:
    candidates = [p for p in WORKDIR.glob("*.xlsm") if p.name != "sales_hub.xlsm"]
    if not candidates:
        raise FileNotFoundError("No xlsm workbook found in workspace.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def month_sheet_names(wb) -> list[str]:
    names = [
        name
        for name in wb.sheetnames
        if name.endswith(MONTH_SUFFIX)
        and not name.startswith(RAW_PREFIX)
        and name[:-1].isdigit()
    ]
    names.sort(key=lambda x: int(x[:-1]))
    return names


def excluded_item(label: str) -> bool:
    return any(term in label for term in EXCLUDED_TERMS)


def is_summary_row(item_label: str) -> bool:
    return (
        not item_label
        or item_label == TOTAL_LABEL
        or item_label == TA_TOTAL_LABEL
        or item_label.endswith(SUBTOTAL_SUFFIX)
    )


def month_items(ws) -> list[dict]:
    items: list[dict] = []
    current_line = ""

    for row in range(5, ws.max_row + 1):
        line_value = ws.cell(row, 1).value
        item_value = ws.cell(row, 2).value

        if line_value:
            current_line = str(line_value).strip()

        item_label = "" if item_value is None else str(item_value).strip()
        if is_summary_row(item_label):
            continue
        if excluded_item(item_label):
            continue

        month_qty = ws.cell(row, 3).value or 0
        prev_qty = ws.cell(row, 4).value or 0
        month_amt = ws.cell(row, 7).value or 0
        prev_amt = ws.cell(row, 8).value or 0
        cum_qty = ws.cell(row, 11).value or 0
        cum_prev_qty = ws.cell(row, 12).value or 0
        cum_amt = ws.cell(row, 15).value or 0
        cum_prev_amt = ws.cell(row, 16).value or 0

        items.append(
            {
                "line": current_line,
                "item": item_label,
                "month_qty": float(month_qty),
                "prev_qty": float(prev_qty),
                "month_amt": float(month_amt),
                "prev_amt": float(prev_amt),
                "cum_qty": float(cum_qty),
                "cum_prev_qty": float(cum_prev_qty),
                "cum_amt": float(cum_amt),
                "cum_prev_amt": float(cum_prev_amt),
            }
        )

    return items


def pct(curr: float, prev: float) -> float:
    if prev == 0:
        return 0.0 if curr == 0 else 1.0
    return (curr / prev) - 1


def style_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, *, fill=None, bold=False, center=False) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            if fill is not None:
                cell.fill = fill
            if bold:
                cell.font = Font(bold=True)
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def write_monthly_overview(ws, month_data: dict[str, list[dict]], start_row: int) -> int:
    ws.cell(start_row, 1).value = "\uC6D4\uBCC4 \uC804\uCCB4 \uC694\uC57D"
    ws.cell(start_row, 1).font = Font(bold=True, size=12)
    start_row += 1

    headers = [
        "\uC6D4",
        "\uB2F9\uC6D4\uC218\uB7C9",
        "\uC804\uB144\uC218\uB7C9",
        "\uC99D\uAC10",
        "\uB2F9\uC6D4\uAE08\uC561",
        "\uC804\uB144\uAE08\uC561",
        "YOY",
        "\uB204\uACC4\uC218\uB7C9",
        "\uC804\uB144\uB204\uACC4\uC218\uB7C9",
        "\uB204\uACC4\uC99D\uAC10",
        "\uB204\uACC4\uAE08\uC561",
        "\uC804\uB144\uB204\uACC4\uAE08\uC561",
        "\uB204\uACC4YOY",
    ]
    for idx, text in enumerate(headers, start=1):
        ws.cell(start_row, idx).value = text
    style_range(ws, start_row, start_row, 1, 13, fill=HEADER_FILL, bold=True, center=True)

    row = start_row + 1
    for month_name, items in month_data.items():
        month_qty = sum(x["month_qty"] for x in items)
        prev_qty = sum(x["prev_qty"] for x in items)
        month_amt = sum(x["month_amt"] for x in items)
        prev_amt = sum(x["prev_amt"] for x in items)
        cum_qty = sum(x["cum_qty"] for x in items)
        cum_prev_qty = sum(x["cum_prev_qty"] for x in items)
        cum_amt = sum(x["cum_amt"] for x in items)
        cum_prev_amt = sum(x["cum_prev_amt"] for x in items)

        values = [
            month_name,
            month_qty,
            prev_qty,
            pct(month_qty, prev_qty),
            month_amt,
            prev_amt,
            pct(month_amt, prev_amt),
            cum_qty,
            cum_prev_qty,
            pct(cum_qty, cum_prev_qty),
            cum_amt,
            cum_prev_amt,
            pct(cum_amt, cum_prev_amt),
        ]
        for idx, value in enumerate(values, start=1):
            ws.cell(row, idx).value = value
        row += 1

    style_range(ws, start_row + 1, row - 1, 1, 13)
    for r in range(start_row + 1, row):
        for col in (2, 3, 8, 9):
            ws.cell(r, col).number_format = "#,##0"
        for col in (5, 6, 11, 12):
            ws.cell(r, col).number_format = '"\u20A9"#,##0'
        for col in (4, 7, 10, 13):
            ws.cell(r, col).number_format = "0.0%"

    return row + 2


def write_line_sections(ws, month_data: dict[str, list[dict]], start_row: int) -> int:
    for month_name, items in month_data.items():
        ws.cell(start_row, 1).value = f"{month_name} \uB77C\uC778\uBCC4 \uC694\uC57D"
        ws.cell(start_row, 1).font = Font(bold=True, size=12)
        start_row += 1

        headers = [
            "\uB77C\uC778",
            "\uB2F9\uC6D4\uC218\uB7C9",
            "\uC804\uB144\uC218\uB7C9",
            "\uC99D\uAC10",
            "\uB2F9\uC6D4\uAE08\uC561",
            "\uC804\uB144\uAE08\uC561",
            "YOY",
        ]
        for idx, text in enumerate(headers, start=1):
            ws.cell(start_row, idx).value = text
        style_range(ws, start_row, start_row, 1, 7, fill=HEADER_FILL, bold=True, center=True)

        grouped: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
        for item in items:
            line = item["line"] or "\uAE30\uD0C0"
            grouped[line]["month_qty"] += item["month_qty"]
            grouped[line]["prev_qty"] += item["prev_qty"]
            grouped[line]["month_amt"] += item["month_amt"]
            grouped[line]["prev_amt"] += item["prev_amt"]

        row = start_row + 1
        for line in grouped:
            month_qty = grouped[line]["month_qty"]
            prev_qty = grouped[line]["prev_qty"]
            month_amt = grouped[line]["month_amt"]
            prev_amt = grouped[line]["prev_amt"]
            values = [
                line,
                month_qty,
                prev_qty,
                pct(month_qty, prev_qty),
                month_amt,
                prev_amt,
                pct(month_amt, prev_amt),
            ]
            for idx, value in enumerate(values, start=1):
                ws.cell(row, idx).value = value
            row += 1

        style_range(ws, start_row + 1, row - 1, 1, 7)
        for r in range(start_row + 1, row):
            for col in (2, 3):
                ws.cell(r, col).number_format = "#,##0"
            for col in (5, 6):
                ws.cell(r, col).number_format = '"\u20A9"#,##0'
            for col in (4, 7):
                ws.cell(r, col).number_format = "0.0%"

        start_row = row + 2

    return start_row


def write_top10_sections(ws, month_data: dict[str, list[dict]], start_row: int) -> int:
    for month_name, items in month_data.items():
        ws.cell(start_row, 1).value = f"{month_name} \uC0C1\uC704 \uD488\uBAA9 TOP10"
        ws.cell(start_row, 1).font = Font(bold=True, size=12)
        start_row += 1

        headers = [
            "\uC21C\uC704",
            "\uB77C\uC778",
            "\uBCF4\uACE0\uC0C1\uD488\uBA85",
            "\uB2F9\uC6D4\uC218\uB7C9",
            "\uB2F9\uC6D4\uAE08\uC561",
        ]
        for idx, text in enumerate(headers, start=1):
            ws.cell(start_row, idx).value = text
        style_range(ws, start_row, start_row, 1, 5, fill=HEADER_FILL, bold=True, center=True)

        ranked = sorted(items, key=lambda x: abs(x["month_amt"]), reverse=True)[:10]
        row = start_row + 1
        for idx, item in enumerate(ranked, start=1):
            ws.cell(row, 1).value = idx
            ws.cell(row, 2).value = item["line"]
            ws.cell(row, 3).value = item["item"]
            ws.cell(row, 4).value = item["month_qty"]
            ws.cell(row, 5).value = item["month_amt"]
            row += 1

        if ranked:
            style_range(ws, start_row + 1, row - 1, 1, 5)
            for r in range(start_row + 1, row):
                ws.cell(r, 4).number_format = "#,##0"
                ws.cell(r, 5).number_format = '"\u20A9"#,##0'

        start_row = row + 2

    return start_row


def apply_layout(ws) -> None:
    widths = {
        "A": 16,
        "B": 16,
        "C": 28,
        "D": 14,
        "E": 16,
        "F": 16,
        "G": 12,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 16,
        "L": 16,
        "M": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 13) + 1):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left" if col == 3 else "center",
            )


def main() -> None:
    xlsm = latest_program_file()
    backup = xlsm.with_name(f"{xlsm.stem}_backup_summary.xlsm")
    copy2(xlsm, backup)

    wb = load_workbook(xlsm, keep_vba=True)
    month_names = month_sheet_names(wb)
    month_data = {name: month_items(wb[name]) for name in month_names}

    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]

    ws = wb.create_sheet(SUMMARY_SHEET, 1)
    ws["A1"] = "\uC6D4\uAC04 \uC694\uC57D \uBCF4\uACE0"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:M1")
    style_range(ws, 1, 1, 1, 13, fill=TITLE_FILL)

    next_row = 3
    next_row = write_monthly_overview(ws, month_data, next_row)
    next_row = write_line_sections(ws, month_data, next_row)
    write_top10_sections(ws, month_data, next_row)
    apply_layout(ws)

    wb.save(xlsm)
    wb.close()


if __name__ == "__main__":
    main()
