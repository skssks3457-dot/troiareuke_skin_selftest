from __future__ import annotations

import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


INPUT_PATH = Path(r"C:\Users\MK\Desktop\전사 품목별 판매 집계 허브.xlsx")
OPTIONAL_PRICE_PATHS = [
    Path(r"C:\Users\MK\Desktop\20260406_밀리맘 상품 판매가 리스트.xlsx"),
    Path(r"C:\Users\MK\Desktop\(공유용) 20260406 기준 상품별 소비자가.xlsx"),
]
OLD_HUB_PATH = Path(r"C:\Users\MK\Desktop\바탕화면 1차\자동화\채널별 상품 실적 집계 허브.xlsm")
OUTPUT_PATH = Path(r"C:\Users\MK\Documents\Playground\판매분석_전처리_기초.xlsx")
RETAIL_DEPTS = ["트로이아르케사업팀", "셀럽B2B", "밀리맘사업팀"]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass
class SourceRow:
    row_no: int
    code: str
    name: str
    qty: float
    total: float
    line: str
    division: str
    composition: float | None
    sold_qty_hint: float | None
    retail_price: float | None
    shop_price: float | None


def infer_division(row: SourceRow) -> tuple[str, str]:
    name = row.name.upper()
    if row.division in ("단독", "단품"):
        if any(token in name for token in ["1+1", "2+1", "3+1", "4+1", "5+1", "6+6", "3+3", "2개", "3개", "SET", "세트", "키트", "기프트", "+"]):
            return "세트", "기존 구분 보정"
        return "단품", "기존 구분 사용"
    if row.division == "세트":
        return "세트", "기존 구분 사용"

    if any(token in name for token in ["1+1", "2+1", "3+1", "4+1", "5+1", "6+6", "3+3", "2+2"]):
        return "세트", "프로모션 패턴"
    if any(token in name for token in ["세트", "SET", "키트", "KIT", "기프트", "GIFT", "+"]):
        return "세트", "세트 키워드"
    if re.search(r"(\d+)\s*개", row.name):
        return "세트", "묶음 수량 패턴"
    if re.search(r"(\d+)\s*(?:SET|세트)", row.name, flags=re.IGNORECASE):
        return "세트", "세트 수량 패턴"
    return "단품", "기본값"


@dataclass
class RuleRow:
    set_code: str
    set_name: str
    line: str
    rule_type: str
    component_code: str
    component_name: str
    component_physical_qty: float
    component_settle_qty: float
    allocation_ratio: float
    match_status: str
    note: str


@dataclass
class OldHubReference:
    product_prices: dict[str, dict[str, float | str | None]]
    set_rules: dict[str, list[RuleRow]]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def load_source_rows(path: Path) -> list[SourceRow]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[SourceRow] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code = clean_text(row[0])
        name = clean_text(row[1])
        if not code and not name:
            continue
        rows.append(
            SourceRow(
                row_no=row_no,
                code=code,
                name=name,
                qty=as_number(row[2]) or 0.0,
                total=as_number(row[3]) or 0.0,
                line=clean_text(row[4]),
                division=clean_text(row[5]),
                composition=as_number(row[6]),
                sold_qty_hint=as_number(row[7]),
                retail_price=as_number(row[8]),
                shop_price=as_number(row[9]),
            )
        )
    return rows


def load_reference_prices(paths: Iterable[Path]) -> dict[str, dict[str, float | str | None]]:
    price_map: dict[str, dict[str, float | str | None]] = {}
    for path in paths:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            if any(clean_text(v) == "품목코드" for v in first_row):
                for row in ws.iter_rows(min_row=2, values_only=True):
                    code = clean_text(row[0]) if len(row) > 0 else ""
                    if not code:
                        continue
                    price_map[code] = {
                        "상품명": clean_text(row[1]) if len(row) > 1 else "",
                        "소비자가": as_number(row[5]) if len(row) > 5 else None,
                        "샵가": None,
                        "가격출처": path.name,
                        "카테고리": clean_text(row[2]) if len(row) > 2 else "",
                    }
                continue
            current_category = ""
            for row in ws.iter_rows(values_only=True):
                col_a = clean_text(row[0]) if len(row) > 0 else ""
                code = clean_text(row[14]) if len(row) > 14 else ""
                name = clean_text(row[15]) if len(row) > 15 else ""
                retail = as_number(row[16]) if len(row) > 16 else None
                if col_a and not code:
                    current_category = col_a
                if not code:
                    continue
                price_map[code] = {
                    "상품명": name,
                    "소비자가": retail,
                    "샵가": None,
                    "가격출처": path.name,
                    "카테고리": current_category,
                }
    return price_map


def load_old_hub_reference(path: Path) -> OldHubReference:
    if not path.exists():
        return OldHubReference(product_prices={}, set_rules={})

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    product_prices: dict[str, dict[str, float | str | None]] = {}
    set_rules: dict[str, list[RuleRow]] = defaultdict(list)

    for ws in wb.worksheets:
        first_row = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        if first_row[:2] == ["품목코드", "ERP 기준 품목명"]:
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = clean_text(row[0])
                if not code:
                    continue
                product_prices[code] = {
                    "상품명": clean_text(row[2]) or clean_text(row[1]),
                    "소비자가": as_number(row[5]),
                    "샵가": None,
                    "가격출처": f"{path.name}:단품상품리스트",
                    "카테고리": clean_text(row[4]),
                }
        elif first_row[:2] == ["세트코드", "세트상품"]:
            for row in ws.iter_rows(min_row=2, values_only=True):
                set_code = clean_text(row[0])
                component_code = clean_text(row[2])
                if not set_code or not component_code:
                    continue
                rule = RuleRow(
                    set_code=set_code,
                    set_name=clean_text(row[1]),
                    line="",
                    rule_type="과거허브 매핑",
                    component_code=component_code,
                    component_name=clean_text(row[3]),
                    component_physical_qty=as_number(row[4]) or 1.0,
                    component_settle_qty=as_number(row[5]) or as_number(row[4]) or 1.0,
                    allocation_ratio=0.0,
                    match_status="과거매핑",
                    note="과거 허브 세트코드 매핑 재사용",
                )
                set_rules[set_code].append(rule)

    for set_code, rule_rows in set_rules.items():
        total_settle = sum(rule.component_settle_qty for rule in rule_rows) or 1.0
        for rule in rule_rows:
            rule.allocation_ratio = rule.component_settle_qty / total_settle

    return OldHubReference(product_prices=product_prices, set_rules=set_rules)


def normalize_name(name: str) -> str:
    text = name.upper()
    text = re.sub(r"\(사용중단\)", "", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\[[^\]]*\]", " ", text)
    text = re.sub(r"1\+1|2\+1|3\+1|4\+1", " ", text)
    text = re.sub(r"\bSET\b|\bKIT\b|\bNEW\b|\bP\b|세트|기프트|키트", " ", text)
    text = re.sub(r"[0-9]+ML|[0-9]+G|[0-9]+EA|[0-9]+개", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"[/,+*&\-]", " ", text)
    text = re.sub(r"\s+", "", text)
    return text


def infer_same_item_counts(name: str, composition: float | None, sold_qty_hint: float | None) -> tuple[float | None, float | None, str]:
    if composition and sold_qty_hint:
        return composition, sold_qty_hint, "기존입력값 활용"
    promo = re.search(r"(\d+)\s*\+\s*(\d+)", name)
    if promo:
        paid = float(promo.group(1))
        free = float(promo.group(2))
        return paid + free, paid, f"{int(paid)}+{int(free)} 자동인식"
    multi = re.search(r"(\d+)\s*(?:SET|set)", name)
    if multi:
        qty = float(multi.group(1))
        return qty, qty, f"{int(qty)}SET 자동인식"
    korean_multi = re.search(r"(\d+)\s*세트", name)
    if korean_multi:
        qty = float(korean_multi.group(1))
        return qty, qty, f"{int(qty)}세트 자동인식"
    repeated = re.search(r"\[(?:[^\]]*?)(\d+)\s*EA[^\]]*\]", name, flags=re.IGNORECASE)
    if repeated:
        qty = float(repeated.group(1))
        return qty, qty, "ea 자동인식"
    return None, None, ""


def strip_set_suffix(code: str) -> str:
    text = code
    text = re.sub(r"/\d+$", "", text)
    text = re.sub(r"-(?:P\d+)$", "", text)
    text = re.sub(r"-(?:\d+|[A-Z]+\d*)$", "", text)
    return text


def is_set_like(row: SourceRow) -> bool:
    division, _ = infer_division(row)
    return division == "세트"


def split_bracket_components(name: str) -> list[tuple[str, float]]:
    match = re.search(r"\[([^\]]+)\]", name)
    if not match:
        return []
    raw = match.group(1)
    if re.fullmatch(r"\s*\d+\s*종\s*", raw):
        return []
    parts = [part.strip() for part in re.split(r"\s*\+\s*", raw) if part.strip()]
    results: list[tuple[str, float]] = []
    for part in parts:
        qty = 1.0
        patterns = [
            r"(\d+)\s*EA",
            r"(\d+)\s*ea",
            r"\*\s*(\d+)",
            r"(\d+)\s*개",
            r"(\d+)\s*SET",
            r"(\d+)$",
        ]
        for pattern in patterns:
            found = re.search(pattern, part, flags=re.IGNORECASE)
            if found:
                qty = float(found.group(1))
                part = re.sub(pattern, "", part, flags=re.IGNORECASE).strip(" */")
                break
        part = re.sub(r"\s+", " ", part).strip()
        if part:
            results.append((part, qty))
    return results


def find_best_name_match(component_name: str, product_lookup: dict[str, SourceRow]) -> tuple[str, str, str]:
    if not component_name:
        return "", "", "검토필요"
    comp_key = normalize_name(component_name)
    best_code = ""
    best_name = ""
    best_score = 0.0
    for code, row in product_lookup.items():
        product_key = normalize_name(row.name)
        if not product_key:
            continue
        if comp_key and (comp_key in product_key or product_key in comp_key):
            score = min(len(comp_key), len(product_key)) / max(len(comp_key), len(product_key))
        else:
            overlap = len(set(re.findall(r".", comp_key)) & set(re.findall(r".", product_key)))
            score = overlap / max(len(set(product_key)), 1)
        if score > best_score:
            best_score = score
            best_code = code
            best_name = row.name
    if best_score >= 0.7:
        return best_code, best_name, "자동매칭"
    return "", component_name, "검토필요"


def find_best_product_by_set_name(set_name: str, product_lookup: dict[str, SourceRow]) -> tuple[str, str, float]:
    set_key = normalize_name(set_name)
    best_code = ""
    best_name = ""
    best_score = 0.0
    for code, row in product_lookup.items():
        product_key = normalize_name(row.name)
        if not product_key:
            continue
        if product_key and product_key in set_key:
            score = len(product_key) / max(len(set_key), 1)
        else:
            overlap = len(set(product_key) & set(set_key))
            score = overlap / max(len(set(product_key)), 1)
        if score > best_score:
            best_code = code
            best_name = row.name
            best_score = score
    return best_code, best_name, best_score


def build_rules(rows: list[SourceRow], old_hub_rules: dict[str, list[RuleRow]]) -> tuple[list[RuleRow], dict[str, list[RuleRow]]]:
    product_lookup = {row.code: row for row in rows if not is_set_like(row)}
    all_codes = set(product_lookup)
    rules: list[RuleRow] = []
    rules_by_set: dict[str, list[RuleRow]] = defaultdict(list)

    for row in rows:
        if not is_set_like(row):
            continue

        if row.code in old_hub_rules:
            copied_rules = []
            for old_rule in old_hub_rules[row.code]:
                copied_rule = RuleRow(
                    set_code=row.code,
                    set_name=row.name,
                    line=row.line,
                    rule_type=old_rule.rule_type,
                    component_code=old_rule.component_code,
                    component_name=old_rule.component_name,
                    component_physical_qty=old_rule.component_physical_qty,
                    component_settle_qty=old_rule.component_settle_qty,
                    allocation_ratio=old_rule.allocation_ratio,
                    match_status=old_rule.match_status,
                    note=old_rule.note,
                )
                copied_rules.append(copied_rule)
                rules.append(copied_rule)
            rules_by_set[row.code].extend(copied_rules)
            continue

        physical_qty, settle_qty, same_note = infer_same_item_counts(row.name, row.composition, row.sold_qty_hint)
        base_code = strip_set_suffix(row.code)
        base_exists = base_code in all_codes
        matched_code = base_code
        matched_name = product_lookup[base_code].name if base_exists else ""

        if physical_qty and settle_qty and not base_exists:
            guessed_code, guessed_name, guessed_score = find_best_product_by_set_name(row.name, product_lookup)
            if guessed_score >= 0.72:
                matched_code = guessed_code
                matched_name = guessed_name
                base_exists = True
                same_note = (same_note + " / 이름유사도매칭").strip(" /")

        if physical_qty and settle_qty and base_exists:
            rule = RuleRow(
                set_code=row.code,
                set_name=row.name,
                line=row.line,
                rule_type="동일상품 프로모션",
                component_code=matched_code,
                component_name=matched_name,
                component_physical_qty=physical_qty,
                component_settle_qty=settle_qty,
                allocation_ratio=1.0,
                match_status="자동완료",
                note=same_note,
            )
            rules.append(rule)
            rules_by_set[row.code].append(rule)
            continue

        components = split_bracket_components(row.name)
        if components:
            ratio = round(1 / len(components), 6)
            for component_name, qty in components:
                comp_code, comp_name, status = find_best_name_match(component_name, product_lookup)
                rule = RuleRow(
                    set_code=row.code,
                    set_name=row.name,
                    line=row.line,
                    rule_type="혼합세트 균등배분",
                    component_code=comp_code,
                    component_name=comp_name,
                    component_physical_qty=qty,
                    component_settle_qty=ratio,
                    allocation_ratio=ratio,
                    match_status=status,
                    note="혼합세트는 기본적으로 균등배분. 필요시 비율 조정",
                )
                rules.append(rule)
                rules_by_set[row.code].append(rule)
            continue

        rule = RuleRow(
            set_code=row.code,
            set_name=row.name,
            line=row.line,
            rule_type="수동정의 필요",
            component_code="",
            component_name="",
            component_physical_qty=1.0,
            component_settle_qty=1.0,
            allocation_ratio=1.0,
            match_status="검토필요",
            note="구성요소 자동인식 실패",
        )
        rules.append(rule)
        rules_by_set[row.code].append(rule)

    return rules, rules_by_set


def auto_width(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        max_len = 0
        for cell in column:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def style_header(ws, row_no: int = 1) -> None:
    for cell in ws[row_no]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def merge_price_maps(*maps: dict[str, dict[str, float | str | None]]) -> dict[str, dict[str, float | str | None]]:
    merged: dict[str, dict[str, float | str | None]] = {}
    for price_map in maps:
        for code, meta in price_map.items():
            current = merged.get(code, {}).copy()
            current.update({k: v for k, v in meta.items() if v not in (None, "")})
            merged[code] = current
    for meta in merged.values():
        retail = meta.get("소비자가")
        if retail not in (None, "") and meta.get("샵가") in (None, ""):
            meta["샵가"] = round(float(retail) * 0.5, 2)
    return merged


def create_output(rows: list[SourceRow], price_map: dict[str, dict[str, float | str | None]], rules: list[RuleRow], rules_by_set: dict[str, list[RuleRow]]) -> None:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "사용안내"
    guide_rows = [
        ("시트", "설명"),
        ("상품마스터", "가격을 직접 보완하는 기준 시트. 소비자가/샵가/구분을 여기서 관리"),
        ("세트규칙", "세트를 단품으로 풀어낼 기준표. 필요한 세트만 수정"),
        ("체크포인트", "세트규칙에서 아직 손봐야 하는 세트만 모아둔 시트"),
    ]
    for row in guide_rows:
        ws_guide.append(row)
    style_header(ws_guide)

    ws_master = wb.create_sheet("상품마스터")
    master_header = [
        "품목코드",
        "품목명",
        "라인",
        "구분",
        "구분판단",
        "현재평균판매단가",
        "소비자가",
        "가격출처",
        "실물수량기준",
        "정산수량기준",
        "검토상태",
        "비고",
    ]
    ws_master.append(master_header)
    set_codes = set(r.set_code for r in rules)
    for row in rows:
        avg_price = row.total / row.qty if row.qty else None
        price_meta = price_map.get(row.code, {})
        inferred_division, division_reason = infer_division(row)
        auto_class = "세트" if row.code in set_codes else inferred_division
        physical_qty, settle_qty, note = infer_same_item_counts(row.name, row.composition, row.sold_qty_hint)
        review = "가격검토" if auto_class != "세트" and not price_meta.get("소비자가") else "확인"
        if row.code in rules_by_set:
            statuses = {r.match_status for r in rules_by_set[row.code]}
            review = "검토필요" if "검토필요" in statuses else "자동완료"
        ws_master.append(
            [
                row.code,
                row.name,
                row.line,
                auto_class,
                division_reason,
                round(avg_price, 2) if avg_price is not None else None,
                price_meta.get("소비자가", row.retail_price),
                price_meta.get("가격출처", "원본없음"),
                physical_qty if physical_qty is not None else (row.composition or 1),
                settle_qty if settle_qty is not None else (row.sold_qty_hint or 1),
                review,
                note,
            ]
        )
    style_header(ws_master)

    ws_rules = wb.create_sheet("세트규칙")
    rules_header = [
        "세트코드",
        "세트명",
        "라인",
        "규칙유형",
        "구성품코드",
        "구성품명",
        "세트1개당 실물수량",
        "세트1개당 정산수량",
        "매출배분비율",
        "매칭상태",
        "비고",
    ]
    ws_rules.append(rules_header)
    for rule in rules:
        ws_rules.append(
            [
                rule.set_code,
                rule.set_name,
                rule.line,
                rule.rule_type,
                rule.component_code,
                rule.component_name,
                rule.component_physical_qty,
                rule.component_settle_qty,
                rule.allocation_ratio,
                rule.match_status,
                rule.note,
            ]
        )
    style_header(ws_rules)

    ws_check = wb.create_sheet("체크포인트")
    ws_check.append(["세트코드", "세트명", "매칭상태", "비고"])
    seen = set()
    for rule in rules:
        if rule.match_status == "검토필요" and rule.set_code not in seen:
            ws_check.append([rule.set_code, rule.set_name, rule.match_status, rule.note])
            seen.add(rule.set_code)
    style_header(ws_check)

    for ws in wb.worksheets:
        auto_width(ws)
        ws.freeze_panes = "A2"

    for row in ws_rules.iter_rows(min_row=2):
        if row[9].value == "검토필요":
            for cell in row:
                cell.fill = WARN_FILL
        elif row[9].value == "자동완료":
            for cell in row:
                cell.fill = GOOD_FILL

    for row in ws_master.iter_rows(min_row=2):
        if row[11].value == "검토필요":
            for cell in row:
                cell.fill = WARN_FILL

    wb.save(OUTPUT_PATH)


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"원본 파일이 없습니다: {INPUT_PATH}")
    rows = load_source_rows(INPUT_PATH)
    price_map = load_reference_prices(OPTIONAL_PRICE_PATHS)
    old_hub = load_old_hub_reference(OLD_HUB_PATH)
    merged_price_map = merge_price_maps(old_hub.product_prices, price_map)
    rules, rules_by_set = build_rules(rows, old_hub.set_rules)
    create_output(rows, merged_price_map, rules, rules_by_set)
    unresolved_sets = len({rule.set_code for rule in rules if rule.match_status == "검토필요"})
    auto_sets = len({rule.set_code for rule in rules if rule.match_status in {"자동완료", "과거매핑"}})
    print(f"완료: {OUTPUT_PATH}")
    print(f"전체 품목 수: {len(rows)}")
    print(f"세트 자동완료: {auto_sets}")
    print(f"세트 검토필요: {unresolved_sets}")


if __name__ == "__main__":
    main()
