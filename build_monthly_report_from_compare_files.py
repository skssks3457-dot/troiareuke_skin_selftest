from __future__ import annotations

import glob
import os
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
DESKTOP = Path(r"C:\Users\MK\Desktop")
SOURCE_PATH = WORKDIR / "판매분석_전처리_기초.xlsx"
OUTPUT_PATH = WORKDIR / "경영회의자료_MK_1~3월_자동집계.xlsx"


REPORT_LAYOUT = [
    ("트로이아르케", "트로이필(웨폰키트)"),
    ("트로이아르케", "VVS키트"),
    ("트로이아르케", "힐링칵테일"),
    ("트로이아르케", "악센 리커버리"),
    ("트로이아르케", "악센 오일컷클렌징"),
    ("트로이아르케", "AGT하이드로 크림"),
    ("트로이아르케", "악센 센앰플"),
    ("트로이아르케", "AGT하이드로 에센스"),
    ("트로이아르케", "에너지크림"),
    ("트로이아르케", "쉴드크림"),
    ("트로이아르케", "아이크림"),
    ("트로이아르케", "악센 AC스팟솔루션"),
    ("트로이아르케", "악센 UV 프로텍터에센스"),
    ("트로이아르케", "인텐스 UV프로텍터크림"),
    ("트로이아르케", "멜라C인퓨저"),
    ("트로이아르케", "GPS마스크 3종"),
    ("트로이아르케", "에스테틱 비비크림"),
    ("트로이아르케", "트로이아르케쿠션"),
    ("트로이아르케", "기타(링플, 샤쉐, LD, 악센)"),
    ("TA 서울라인", "엑토인 톤업 선세럼"),
    ("TA 서울라인", "옥시젠 토닝 클렌저"),
    ("TA 서울라인", "워터인 퀀칭 세럼"),
    ("TA 서울라인", "엑토인 퀀칭 크림"),
    ("TA 서울라인", "더마인 리+베리어"),
    ("TA 서울라인", "신부쿠션 21호/22호"),
    ("TA 서울라인", "기타(세트상품)"),
    ("미용기기", "멜라소닉(부속품포함)"),
    ("미용기기", "아크소닉(부속품포함)"),
    ("미용기기", "엘디소닉(부속품포함)"),
    ("미용기기", "AI피부진단기"),
    ("담 향", "소프트브레스"),
    ("담 향", "슈퍼아로마"),
    ("담 향", "기타"),
    ("밀리맘", "오일앤크림"),
    ("밀리맘", "아토크림"),
    ("밀리맘", "새싹 로션"),
    ("밀리맘", "새싹 워시"),
    ("밀리맘", "기타(젤, 힙, 미니, 세트)"),
    ("기 타", "할인"),
    ("기 타", "뷰티툴(BT)"),
    ("기 타", "파트너스(교육)"),
    ("기 타", "기타(택배비, 리플렛, 단종)"),
]

GROUP_ORDER = []
for g, _ in REPORT_LAYOUT:
    if g not in GROUP_ORDER:
        GROUP_ORDER.append(g)

ITEM_TO_GROUP = {item: group for group, item in REPORT_LAYOUT}


def find_compare_files() -> list[tuple[int, Path]]:
    targets = []
    for month in (1, 2, 3):
        filename = f"{month}월 비교 자료.xlsx"
        path = DESKTOP / filename
        if not path.exists():
            raise FileNotFoundError(f"파일이 없습니다: {path}")
        targets.append((month, path))
    return targets


def load_reference():
    wb = load_workbook(SOURCE_PATH, data_only=False)
    ws_master, ws_rules, ws_map = wb.worksheets[:3]

    master = {}
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        master[str(code).strip()] = {
            "품목명": row[1],
            "라인": row[2],
            "구분": row[3],
            "소비자가": row[4],
            "실물수량기준": row[5] or 1,
            "정산수량기준": row[6] or 1,
        }

    rules = defaultdict(list)
    for row in ws_rules.iter_rows(min_row=2, values_only=True):
        set_code = row[0]
        comp_code = row[3]
        if not set_code or not comp_code:
            continue
        rules[str(set_code).strip()].append(
            {
                "구성품코드": str(comp_code).strip(),
                "구성품명": row[4],
                "실물수량": float(row[5] or 0),
                "정산수량": float(row[6] or 0),
                "배분비율": float(row[7] or 0),
            }
        )

    code_to_report = {}
    for row in ws_map.iter_rows(min_row=2, values_only=True):
        report_name, code = row[0], row[1]
        if report_name and code:
            code_to_report[str(code).strip()] = str(report_name).strip()

    wb.close()
    return master, rules, code_to_report


def parse_month_compare(month: int, path: Path):
    wb = load_workbook(path, data_only=False)
    ws = wb.worksheets[0]
    records = []

    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 3).value
        name = ws.cell(r, 4).value
        if not code or not name:
            continue
        code = str(code).strip()
        name = str(name).strip()

        qty_2026 = ws.cell(r, 5).value
        supply_2026 = ws.cell(r, 6).value
        vat_2026 = ws.cell(r, 7).value
        amt_2026 = ws.cell(r, 8).value
        qty_2025 = ws.cell(r, 9).value
        supply_2025 = ws.cell(r, 10).value
        vat_2025 = ws.cell(r, 11).value
        amt_2025 = ws.cell(r, 12).value

        if any(v is not None for v in [qty_2026, supply_2026, vat_2026, amt_2026]):
            records.append(
                {
                    "기준월": f"2026/{month:02d}",
                    "연도": 2026,
                    "월": month,
                    "품목코드": code,
                    "품목명": name,
                    "수량": safe_num(qty_2026),
                    "공급가액": safe_num(supply_2026),
                    "부가세": safe_num(vat_2026),
                    "합계": safe_num(amt_2026),
                }
            )

        if any(v is not None for v in [qty_2025, supply_2025, vat_2025, amt_2025]):
            records.append(
                {
                    "기준월": f"2025/{month:02d}",
                    "연도": 2025,
                    "월": month,
                    "품목코드": code,
                    "품목명": name,
                    "수량": safe_num(qty_2025),
                    "공급가액": safe_num(supply_2025),
                    "부가세": safe_num(vat_2025),
                    "합계": safe_num(amt_2025),
                }
            )

    wb.close()
    return records


def safe_num(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def normalize_records(raw_records, master, rules, code_to_report):
    normalized = []
    agg = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})

    for rec in raw_records:
        code = rec["품목코드"]
        base_name = rec["품목명"]
        qty = rec["수량"]
        amt = rec["합계"]

        if code in rules:
            for rule in rules[code]:
                norm_code = rule["구성품코드"]
                master_info = master.get(norm_code, {})
                norm_name = master_info.get("품목명") or rule["구성품명"] or norm_code
                line = master_info.get("라인", "")
                div = master_info.get("구분", "단품")
                report_item = code_to_report.get(norm_code, fallback_report_item(norm_name, line, norm_code))
                report_group = ITEM_TO_GROUP.get(report_item, fallback_group(report_item, line))
                norm_qty = qty * rule["정산수량"]
                norm_actual = qty * rule["실물수량"]
                norm_amt = amt * rule["배분비율"]
                normalized.append(
                    {
                        "기준월": rec["기준월"],
                        "연도": rec["연도"],
                        "월": rec["월"],
                        "원본품목코드": code,
                        "원본품목명": base_name,
                        "정규화품목코드": norm_code,
                        "정규화품목명": norm_name,
                        "라인": line,
                        "구분": div,
                        "보고대분류": report_group,
                        "보고상품명": report_item,
                        "원본수량": qty,
                        "실물수량": norm_actual,
                        "정산수량": norm_qty,
                        "비교매출": norm_amt,
                        "원본합계": amt,
                    }
                )
                agg[(rec["연도"], rec["월"], report_group, report_item)]["qty"] += norm_qty
                agg[(rec["연도"], rec["월"], report_group, report_item)]["amt"] += norm_amt
        else:
            master_info = master.get(code, {})
            line = master_info.get("라인", "")
            div = master_info.get("구분", "단품")
            qty_factor = safe_num(master_info.get("정산수량기준", 1))
            actual_factor = safe_num(master_info.get("실물수량기준", 1))
            report_item = code_to_report.get(code, fallback_report_item(base_name, line, code))
            report_group = ITEM_TO_GROUP.get(report_item, fallback_group(report_item, line))
            norm_qty = qty * (qty_factor if qty_factor else 1)
            norm_actual = qty * (actual_factor if actual_factor else 1)
            normalized.append(
                {
                    "기준월": rec["기준월"],
                    "연도": rec["연도"],
                    "월": rec["월"],
                    "원본품목코드": code,
                    "원본품목명": base_name,
                    "정규화품목코드": code,
                    "정규화품목명": master_info.get("품목명") or base_name,
                    "라인": line,
                    "구분": div,
                    "보고대분류": report_group,
                    "보고상품명": report_item,
                    "원본수량": qty,
                    "실물수량": norm_actual,
                    "정산수량": norm_qty,
                    "비교매출": amt,
                    "원본합계": amt,
                }
            )
            agg[(rec["연도"], rec["월"], report_group, report_item)]["qty"] += norm_qty
            agg[(rec["연도"], rec["월"], report_group, report_item)]["amt"] += amt

    return normalized, agg


def fallback_group(report_item: str, line: str) -> str:
    if report_item in ITEM_TO_GROUP:
        return ITEM_TO_GROUP[report_item]
    if "밀리맘" in line:
        return "밀리맘"
    if "서울" in line:
        return "TA 서울라인"
    if "미용기기" in line:
        return "미용기기"
    if "담향" in line:
        return "담 향"
    return "트로이아르케"


def fallback_report_item(name: str, line: str, code: str) -> str:
    text = f"{name} {line} {code}".replace(" ", "")
    if "밀리맘" in line:
        if "오일앤크림" in name:
            return "오일앤크림"
        if "아토크림" in name:
            return "아토크림"
        if "로션" in name:
            return "새싹 로션"
        if "워시" in name:
            return "새싹 워시"
        return "기타(젤, 힙, 미니, 세트)"
    if "서울" in line:
        if "톤업선세럼" in text:
            return "엑토인 톤업 선세럼"
        if "옥시젠토닝클렌저" in text:
            return "옥시젠 토닝 클렌저"
        if "워터인퀀칭세럼" in text:
            return "워터인 퀀칭 세럼"
        if "엑토인퀀칭크림" in text:
            return "엑토인 퀀칭 크림"
        if "더마인" in name or "리+베리어" in name:
            return "더마인 리+베리어"
        if "쿠션" in name:
            return "신부쿠션 21호/22호"
        return "기타(세트상품)"
    if code.startswith("MKBT") or "(BT)" in name:
        return "뷰티툴(BT)"
    if "VVS" in name:
        return "VVS키트"
    if "칵테일" in name:
        return "힐링칵테일"
    if "리커버리" in name:
        return "악센 리커버리"
    if "오일컷클렌징" in name:
        return "악센 오일컷클렌징"
    if "AGT" in name and "크림" in name:
        return "AGT하이드로 크림"
    if "센 앰플" in name or "센앰플" in name:
        return "악센 센앰플"
    if "AGT" in name and "에센스" in name:
        return "AGT하이드로 에센스"
    if "에너지크림" in name:
        return "에너지크림"
    if "쉴드크림" in name:
        return "쉴드크림"
    if "아이크림" in name:
        return "아이크림"
    if "스팟솔루션" in name:
        return "악센 AC스팟솔루션"
    if "UV프로텍터 에센스" in name or "UV프로텍터에센스" in name:
        return "악센 UV 프로텍터에센스"
    if "UV프로텍터크림" in name:
        return "인텐스 UV프로텍터크림"
    if "씨-인퓨저" in name or "씨인퓨저" in text:
        return "멜라C인퓨저"
    if "GPS" in name or "GPS" in code:
        return "GPS마스크 3종"
    if "비비크림" in name:
        return "에스테틱 비비크림"
    if "쿠션" in name:
        return "트로이아르케쿠션"
    if "멜라소닉" in name:
        return "멜라소닉(부속품포함)"
    if "아크소닉" in name:
        return "아크소닉(부속품포함)"
    if "엘디소닉" in name:
        return "엘디소닉(부속품포함)"
    if "AI피부진단" in name:
        return "AI피부진단기"
    if "소프트브레스" in name:
        return "소프트브레스"
    if "슈퍼아로마" in name:
        return "슈퍼아로마"
    if "교육" in name:
        return "파트너스(교육)"
    if "할인" in name:
        return "할인"
    if "택배" in name or "리플렛" in name or "단종" in name:
        return "기타(택배비, 리플렛, 단종)"
    return "기타(링플, 샤쉐, LD, 악센)"


def add_sheet_copy(wb_out, source_ws, title: str):
    ws_new = wb_out.create_sheet(title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = ws_new[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
    for key, dim in source_ws.column_dimensions.items():
        ws_new.column_dimensions[key].width = dim.width
    for key, dim in source_ws.row_dimensions.items():
        ws_new.row_dimensions[key].height = dim.height
    return ws_new


def build_report_sheet(ws, month: int, agg):
    ws["A1"] = "1. 제품별 판매현황"
    ws["A2"] = "구분"
    ws["C2"] = f"{month}월 당월"
    ws["K2"] = f"{month}월 누계"
    headers = ["수량", "전년", "증감", "구성비", "금액", "전년", "YOY", "구성비", "수량", "전년", "YOY", "구성비", "금액", "전년", "YOY", "구성비"]
    for idx, header in enumerate(headers, start=3):
        ws.cell(3, idx).value = header

    fill = PatternFill("solid", fgColor="F2EDE1")
    for row in range(1, 4):
        for col in range(1, 19):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 4
    total_month_qty = 0.0
    total_month_amt = 0.0
    total_cum_qty = 0.0
    total_cum_amt = 0.0

    totals_by_item = []
    report_value_rows = []
    for group, item in REPORT_LAYOUT:
        cur_qty = agg[(2026, month, group, item)]["qty"]
        prev_qty = agg[(2025, month, group, item)]["qty"]
        cur_amt = agg[(2026, month, group, item)]["amt"]
        prev_amt = agg[(2025, month, group, item)]["amt"]
        cur_cum_qty = sum(agg[(2026, m, group, item)]["qty"] for m in range(1, month + 1))
        prev_cum_qty = sum(agg[(2025, m, group, item)]["qty"] for m in range(1, month + 1))
        cur_cum_amt = sum(agg[(2026, m, group, item)]["amt"] for m in range(1, month + 1))
        prev_cum_amt = sum(agg[(2025, m, group, item)]["amt"] for m in range(1, month + 1))
        totals_by_item.append((group, item, cur_qty, prev_qty, cur_amt, prev_amt, cur_cum_qty, prev_cum_qty, cur_cum_amt, prev_cum_amt))
        total_month_qty += cur_qty
        total_month_amt += cur_amt
        total_cum_qty += cur_cum_qty
        total_cum_amt += cur_cum_amt

    current_group = None
    group_rows = []
    for idx, entry in enumerate(totals_by_item):
        group, item, cur_qty, prev_qty, cur_amt, prev_amt, cur_cum_qty, prev_cum_qty, cur_cum_amt, prev_cum_amt = entry
        if current_group is None:
            current_group = group
        if group != current_group:
            write_subtotal_row(ws, row_idx, group_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt)
            row_idx += 1
            group_rows = []
            current_group = group

        if not group_rows:
            ws.cell(row_idx, 1).value = group
        ws.cell(row_idx, 2).value = item
        values = [
            cur_qty,
            prev_qty,
            yoy(cur_qty, prev_qty),
            share(cur_qty, total_month_qty),
            cur_amt,
            prev_amt,
            yoy(cur_amt, prev_amt),
            share(cur_amt, total_month_amt),
            cur_cum_qty,
            prev_cum_qty,
            yoy(cur_cum_qty, prev_cum_qty),
            share(cur_cum_qty, total_cum_qty),
            cur_cum_amt,
            prev_cum_amt,
            yoy(cur_cum_amt, prev_cum_amt),
            share(cur_cum_amt, total_cum_amt),
        ]
        report_value_rows.append(values)
        for col, value in enumerate(values, start=3):
            ws.cell(row_idx, col).value = value
        group_rows.append(values)
        row_idx += 1

    if group_rows:
        write_subtotal_row(ws, row_idx, group_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt)
        row_idx += 1

    write_subtotal_row(
        ws,
        row_idx,
        report_value_rows,
        total_month_qty,
        total_month_amt,
        total_cum_qty,
        total_cum_amt,
        grand=True,
    )

    format_report_sheet(ws, row_idx)


def write_subtotal_row(ws, row_idx, value_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt, grand=False):
    sum_cols = [sum(row[i] for row in value_rows if isinstance(row[i], (int, float))) for i in range(16)]
    out = [
        sum_cols[0],
        sum_cols[1],
        yoy(sum_cols[0], sum_cols[1]),
        share(sum_cols[0], total_month_qty),
        sum_cols[4],
        sum_cols[5],
        yoy(sum_cols[4], sum_cols[5]),
        share(sum_cols[4], total_month_amt),
        sum_cols[8],
        sum_cols[9],
        yoy(sum_cols[8], sum_cols[9]),
        share(sum_cols[8], total_cum_qty),
        sum_cols[12],
        sum_cols[13],
        yoy(sum_cols[12], sum_cols[13]),
        share(sum_cols[12], total_cum_amt),
    ]
    ws.cell(row_idx, 1).value = "합계" if grand else None
    ws.cell(row_idx, 2).value = "합계" if grand else "소계"
    for col, value in enumerate(out, start=3):
        ws.cell(row_idx, col).value = value
    fill_color = "D9E1F2" if grand else "EAF1DD"
    fill = PatternFill("solid", fgColor=fill_color)
    for col in range(1, 19):
        ws.cell(row_idx, col).fill = fill
        ws.cell(row_idx, col).font = Font(bold=True)


def yoy(current, prev):
    if not prev:
        return None
    return (current - prev) / prev


def share(current, total):
    if not total:
        return 0
    return current / total


def format_report_sheet(ws, last_row):
    widths = {"A": 14, "B": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in "CDEFGHIJKLMNOPQR":
        ws.column_dimensions[col].width = 11

    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=18):
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="right" if cell.column >= 3 else "left")

    percent_cols = [5, 6, 9, 10, 13, 14, 17, 18]
    for row in range(4, last_row + 1):
        for col in percent_cols:
            ws.cell(row, col).number_format = "0.0%"
    for row in range(4, last_row + 1):
        for col in [3, 4, 7, 8, 11, 12, 15, 16]:
            ws.cell(row, col).number_format = "#,##0"
    for row in range(1, last_row + 1):
        for col in range(1, 19):
            ws.cell(row, col).border = copy(ws.cell(1, 1).border)


def build_output():
    compare_files = find_compare_files()
    master, rules, code_to_report = load_reference()
    raw_records = []
    raw_by_month = {}
    for month, path in compare_files:
        rows = parse_month_compare(month, path)
        raw_records.extend(rows)
        raw_by_month[month] = rows

    normalized, agg = normalize_records(raw_records, master, rules, code_to_report)

    src_wb = load_workbook(SOURCE_PATH, data_only=False)
    wb_out = Workbook()
    default_ws = wb_out.active
    wb_out.remove(default_ws)

    for ws in src_wb.worksheets:
        add_sheet_copy(wb_out, ws, ws.title)

    ws_info = wb_out.create_sheet("사용안내", 0)
    ws_info["A1"] = "월별 자동집계 결과"
    ws_info["A1"].font = Font(size=18, bold=True)
    ws_info["A3"] = "가정"
    ws_info["A3"].font = Font(bold=True)
    notes = [
        "1. 이번 결과는 1월~3월 비교 자료의 품목코드/품목명 기준으로 집계했습니다.",
        "2. 원본 파일에 부서 컬럼이 없어서 세트는 전체 수량/금액 기준으로 해체했습니다.",
        "3. 세트 수량은 세트규칙의 정산수량, 매출은 매출배분비율을 적용했습니다.",
        "4. 보고상품명은 최신 보고상품명코드매핑 시트를 기준으로 묶었습니다.",
    ]
    for idx, note in enumerate(notes, start=4):
        ws_info.cell(idx, 1).value = note
    ws_info.column_dimensions["A"].width = 100

    for month, rows in raw_by_month.items():
        ws = wb_out.create_sheet(f"원본_{month}월")
        headers = ["기준월", "연도", "월", "품목코드", "품목명", "수량", "공급가액", "부가세", "합계"]
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for col, width in {"A": 10, "B": 8, "C": 6, "D": 18, "E": 52, "F": 10, "G": 12, "H": 10, "I": 12}.items():
            ws.column_dimensions[col].width = width

    ws_norm = wb_out.create_sheet("정규화데이터")
    norm_headers = ["기준월", "연도", "월", "원본품목코드", "원본품목명", "정규화품목코드", "정규화품목명", "라인", "구분", "보고대분류", "보고상품명", "원본수량", "실물수량", "정산수량", "비교매출", "원본합계"]
    ws_norm.append(norm_headers)
    for row in normalized:
        ws_norm.append([row[h] for h in norm_headers])

    for month in (1, 2, 3):
        ws = wb_out.create_sheet(f"{month}월")
        build_report_sheet(ws, month, agg)

    src_wb.close()
    wb_out.save(OUTPUT_PATH)
    wb_out.close()
    return OUTPUT_PATH, len(raw_records), len(normalized)


if __name__ == "__main__":
    out, raw_count, norm_count = build_output()
    print(out)
    print(f"raw_records={raw_count}")
    print(f"normalized_records={norm_count}")
