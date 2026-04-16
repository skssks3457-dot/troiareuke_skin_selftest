from __future__ import annotations

import glob
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
DESKTOP = Path(r"C:\Users\MK\Desktop")
OUTPUT_XLSX = WORKDIR / "판매분석_월별허브.xlsx"


def find_source_workbook() -> Path:
    candidates = []
    for path_str in glob.glob(str(WORKDIR / "*.xlsx")):
        path = Path(path_str)
        if "backup" in path.name.lower() or path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, data_only=False)
            if wb.sheetnames[:3] == ["상품마스터", "세트규칙", "보고상품명코드매핑"]:
                candidates.append((path.stat().st_mtime, path))
            wb.close()
        except Exception:
            continue
    if not candidates:
        raise FileNotFoundError("최신 기준본을 찾지 못했습니다.")
    return sorted(candidates, reverse=True)[0][1]


def find_report_workbook() -> Path:
    for path_str in glob.glob(str(DESKTOP / "*.xlsx")):
        path = Path(path_str)
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, data_only=False)
            if wb.sheetnames[:4] == ["보고", "2월", "3월", "4월"]:
                wb.close()
                return path
            wb.close()
        except Exception:
            continue
    raise FileNotFoundError("보고서 템플릿 파일을 찾지 못했습니다.")


def clear_helper_sheets(wb) -> None:
    for name in ["허브", "DeptRule", "RawData", "NormalizedData", "ReportLayout"]:
        if name in wb.sheetnames:
            ws = wb[name]
            wb.remove(ws)


def build_hub_sheet(wb) -> None:
    ws = wb.create_sheet("허브", 0)
    ws["A1"] = "판매분석 월별 허브"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "사용 순서"
    ws["A3"].font = Font(bold=True)
    steps = [
        '1. "데이터 불러오기" 버튼으로 바탕화면의 월별 원본 파일을 선택합니다.',
        '2. 원본 데이터가 RawData 시트로 복사되고, 자동으로 월별 작년대비 시트가 다시 생성됩니다.',
        '3. 기준 시트는 상품마스터 / 세트규칙 / 보고상품명코드매핑을 그대로 사용합니다.',
        '4. 생성된 1월~12월 시트에서 당월/누계 작년대비를 확인합니다.',
    ]
    for idx, text in enumerate(steps, start=4):
        ws.cell(idx, 1).value = text

    ws["A10"] = "최근 불러온 파일"
    ws["A11"] = "최근 갱신 시각"
    ws["B10"] = "-"
    ws["B11"] = "-"

    ws["D3"] = "버튼 안내"
    ws["D3"].font = Font(bold=True)
    ws["D4"] = "데이터 불러오기"
    ws["D5"] = "월별 시트 다시생성"
    ws["D7"] = "주의"
    ws["D8"] = "원본 파일은 첫 번째 시트의 1행에 아래 헤더가 있어야 합니다."
    ws["D9"] = "기준월 / 연도 / 월 / 거래처그룹2코드 / 거래처그룹2 / 품목코드 / 품목명 / 수량 / 공급가액 / 부가세 / 합계"

    fill = PatternFill("solid", fgColor="F4F1EA")
    for cell in ["A1", "A3", "D3", "D7"]:
        ws[cell].fill = fill

    for col, width in {
        "A": 24,
        "B": 48,
        "C": 4,
        "D": 18,
        "E": 40,
    }.items():
        ws.column_dimensions[col].width = width

    for row in range(1, 15):
        ws.row_dimensions[row].height = 22

    for row in ws.iter_rows(min_row=1, max_row=14, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def build_dept_rule_sheet(wb) -> None:
    ws = wb.create_sheet("DeptRule")
    rows = [
        ["부서명", "비교기준", "세트분해", "매출계산방식", "비고"],
        ["트로이아르케사업팀", "소비자가", "Y", "단품은 소비자가 x 정산수량 / 세트는 원본합계 x 배분비율", "무상출고 제외"],
        ["셀럽B2B", "소비자가", "Y", "단품은 소비자가 x 정산수량 / 세트는 원본합계 x 배분비율", "무상출고 제외"],
        ["밀리맘사업팀", "소비자가", "Y", "단품은 소비자가 x 정산수량 / 세트는 원본합계 x 배분비율", "무상출고 제외"],
        ["기타부서", "원본금액", "N", "원본 수량/금액 그대로 사용", "무상출고 제외"],
    ]
    for row in rows:
        ws.append(row)
    for col, width in {"A": 20, "B": 12, "C": 10, "D": 42, "E": 18}.items():
        ws.column_dimensions[col].width = width
    ws.sheet_state = "hidden"


def build_raw_sheet(wb) -> None:
    ws = wb.create_sheet("RawData")
    headers = ["기준월", "연도", "월", "거래처그룹2코드", "거래처그룹2", "품목코드", "품목명", "수량", "공급가액", "부가세", "합계"]
    ws.append(headers)
    for col, width in {
        "A": 12,
        "B": 8,
        "C": 6,
        "D": 16,
        "E": 22,
        "F": 18,
        "G": 52,
        "H": 10,
        "I": 12,
        "J": 10,
        "K": 12,
    }.items():
        ws.column_dimensions[col].width = width
    ws.sheet_state = "hidden"


def build_normalized_sheet(wb) -> None:
    ws = wb.create_sheet("NormalizedData")
    headers = [
        "기준월",
        "연도",
        "월",
        "부서",
        "원본품목코드",
        "원본품목명",
        "정규화품목코드",
        "정규화품목명",
        "라인",
        "구분",
        "보고대분류",
        "보고상품명",
        "원본수량",
        "실물수량",
        "정산수량",
        "비교매출",
        "원본합계",
    ]
    ws.append(headers)
    for col, width in {
        "A": 12,
        "B": 8,
        "C": 6,
        "D": 18,
        "E": 18,
        "F": 40,
        "G": 18,
        "H": 40,
        "I": 14,
        "J": 8,
        "K": 14,
        "L": 24,
        "M": 10,
        "N": 10,
        "O": 10,
        "P": 14,
        "Q": 14,
    }.items():
        ws.column_dimensions[col].width = width
    ws.sheet_state = "hidden"


def build_report_layout_sheet(wb, report_path: Path) -> None:
    report_wb = load_workbook(report_path, data_only=False)
    report_ws = report_wb["3월"]
    ws = wb.create_sheet("ReportLayout")
    ws.append(["순번", "대분류", "보고상품명"])

    current_group = None
    seq = 1
    for row in range(4, 60):
        group_val = report_ws.cell(row, 1).value
        item_val = report_ws.cell(row, 2).value
        if group_val:
            current_group = str(group_val).strip()
        if not item_val:
            continue
        item_text = str(item_val).strip()
        if item_text == "소계":
            continue
        ws.append([seq, current_group, item_text])
        seq += 1

    report_wb.close()
    for col, width in {"A": 8, "B": 16, "C": 28}.items():
        ws.column_dimensions[col].width = width
    ws.sheet_state = "hidden"


def build_workbook() -> Path:
    source_path = find_source_workbook()
    report_path = find_report_workbook()

    wb = load_workbook(source_path)
    clear_helper_sheets(wb)
    build_hub_sheet(wb)
    build_dept_rule_sheet(wb)
    build_raw_sheet(wb)
    build_normalized_sheet(wb)
    build_report_layout_sheet(wb, report_path)

    wb.save(OUTPUT_XLSX)
    wb.close()
    return OUTPUT_XLSX


if __name__ == "__main__":
    out = build_workbook()
    print(out)
