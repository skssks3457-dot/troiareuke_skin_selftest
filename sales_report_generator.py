from __future__ import annotations

import argparse
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

import openpyxl
import pandas as pd


YEAR26_QTY = ("2026/03/01  ~ 2026/03/31 ", "수량")
YEAR25_QTY = ("2025/03/01  ~ 2025/03/31 ", "수량")


def clean_name(value: str) -> str:
    text = str(value or "").strip()
    text = text.replace("™", "").replace("NEW", "").replace("리뉴얼", "")
    text = re.sub(r"\s+", " ", text)
    return text


def count_pack(text: str, default: int = 1) -> int:
    text = clean_name(text)

    patterns = [r"(\d+)개", r"(\d+)세트", r"(\d+)매입", r"(\d+)회분", r"\*(\d+)(?:개|ea|set)"]

    counts: list[int] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            try:
                counts.append(int(match.group(1)))
            except ValueError:
                continue

    plus_match = re.search(r"(\d+)\s*\+\s*(\d+)", text)
    if plus_match:
        counts.append(int(plus_match.group(1)) + int(plus_match.group(2)))

    if counts:
        return max(counts)
    return default


def count_multiple_keywords(text: str, keywords: list[str]) -> int:
    total = 0
    lowered = clean_name(text)
    for keyword in keywords:
        total += lowered.count(keyword)
    return total


def add_if(mapping: dict[str, int], key: str, amount: int) -> dict[str, int]:
    if amount:
        mapping[key] = mapping.get(key, 0) + amount
    return mapping


def parse_ta_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    result: dict[str, int] = {}

    if "웨폰키트" in text or "트로이필" in text:
        return {"트로이필(웨폰키트)": 1}

    if "VVS 키트" in text:
        return {"VVS키트": max(count_pack(text, 1), 1)}
    if "VVS크림" in text:
        return {"VVS키트": 1}

    if "H+ 칵테일" in text or "힐링칵테일" in text:
        if "그린그린" in text or "블루블루" in text or "레드레드" in text or "엘로우옐로우" in text:
            return {"힐링칵테일": 2}
        if any(word in text for word in ["그린옐로우", "그린레드", "그린블루", "옐로우레드", "옐로우블루", "레드블루"]):
            return {"힐링칵테일": 2}
        result["힐링칵테일"] = count_pack(text, 1)
        if "GPS" in text:
            add_if(result, "GPS마스크 3종", count_multiple_keywords(text, ["GPS"]))
        if "멜라소닉" in text:
            add_if(result, "멜라C인퓨저", count_multiple_keywords(text, ["멜라"]))
        return result

    if "악센 리커버리" in text and "샤쉐" not in text:
        if "봉투" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": 1}
        return {"악센 리커버리": count_pack(text, 1)}

    if "악센 오일컷클렌징" in text and "샤쉐" not in text:
        return {"악센 오일컷클렌징": count_pack(text, 1)}

    if "AGT 하이드로 크림" in text or "AGT하이드로 크림" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        return {"AGT하이드로 크림": count_pack(text, 1)}

    if "센앰플" in text:
        if "미라클 키트" in text:
            return {"악센 센앰플": count_pack(text, 1)}
        return {"악센 센앰플": count_pack(text, 1)}

    if "AGT하이드로 에센스" in text or "AGT 하이드로 에센스" in text:
        if "샤쉐" in text or "AGT듀오" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        return {"AGT하이드로 에센스": count_pack(text, 1)}

    if "에너지크림" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        return {"에너지크림": count_pack(text, 1)}

    if "쉴드크림" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        return {"쉴드크림": count_pack(text, 1)}

    if "안티링클아이크림" in text:
        return {"아이크림": count_pack(text, 1)}

    if "AC스팟솔루션" in text:
        return {"악센 AC스팟솔루션": count_pack(text, 1)}

    if "악센 UV프로텍터에센스" in text or "악센 UV프로텍터 에센스" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        result["악센 UV 프로텍터에센스"] = count_pack(text, 1)
        if "인텐스UV프로텍터크림" in text or "인텐스 UV프로텍터크림" in text or "인텐스 UV프로텍터 크림" in text:
            add_if(result, "인텐스 UV프로텍터크림", 1)
        return result

    if "인텐스 UV프로텍터 크림" in text or "인텐스 UV프로텍터크림" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        result["인텐스 UV프로텍터크림"] = count_pack(text, 1)
        if "악센UV프로텍터에센스" in text or "악센 UV프로텍터에센스" in text or "악센 UV프로텍터 에센스" in text:
            add_if(result, "악센 UV 프로텍터에센스", 1)
        return result

    if "멜라-소닉 씨-인퓨저" in text or "멜라소닉씨인퓨저" in text or "멜라C인퓨저" in text:
        result["멜라C인퓨저"] = count_pack(text, 1)
        if "GPS" in text:
            add_if(result, "GPS마스크 3종", count_multiple_keywords(text, ["GPS"]))
        if "힐칵" in text or "힐링칵테일" in text:
            add_if(result, "힐링칵테일", 1)
        return result

    if "GPS마스크" in text or "GPS 마스크" in text:
        return {"GPS마스크 3종": count_pack(text, 1)}

    if "에스테틱비비크림" in text:
        if "샤쉐" in text:
            return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}
        return {"에스테틱 비비크림": count_multiple_keywords(text, ["비비크림"]) or count_pack(text, 1)}

    if "트로이아르케 쿠션 A+" in text:
        return {"트로이아르케쿠션": count_pack(text, 1)}

    if any(keyword in text for keyword in ["포뮬러 링플", "아로마LD오일", "샤쉐", "트로이얼키트", "스킨레시피", "첫걸음샤쉐", "선케어 2종+비비2종"]):
        return {"기타(링플, 샤쉐, LD, 악센)": count_pack(text, 1)}

    return result


def parse_ts_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    result: dict[str, int] = {}

    if "키트" in text:
        return {"기타(세트상품)": 1}

    if "서울 엑토인 톤업 선세럼" in text:
        add_if(result, "엑토인 톤업 선세럼", count_pack(text, 1))
    if "서울 옥시젠 토닝 클렌저" in text:
        add_if(result, "옥시젠 토닝 클렌저", count_pack(text, 1))
    if "서울 워터인 퀀칭 세럼" in text:
        add_if(result, "워터인 퀀칭 세럼", count_pack(text, 1))
    if "서울 엑토인 퀀칭 크림" in text:
        add_if(result, "엑토인 퀀칭 크림", count_pack(text, 1))
    if "서울 더마인 리 + 베리어" in text or "서울 더마인 리+베리어" in text:
        add_if(result, "더마인 리+베리어", count_pack(text, 1))
    if "서울 에스테틱 쿠션" in text or "서울쿠션" in text:
        add_if(result, "신부쿠션 21호/22호", count_pack(text, 1))

    if result:
        return result
    if text.startswith("(P) 서울 ") or text.startswith("(TS) 서울 ") or "서울 " in text:
        return {"기타(세트상품)": 1}
    return result


def parse_device_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    result: dict[str, int] = {}
    if "멜라소닉" in text:
        return {"멜라소닉(부속품포함)": 1}
    if "아크소닉" in text:
        return {"아크소닉(부속품포함)": 1}
    if "엘디소닉" in text:
        return {"엘디소닉(부속품포함)": 1}
    if "AI피부진단기" in text:
        return {"AI피부진단기": 1}
    return result


def parse_dh_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    if "소프트브레스" in text:
        return {"소프트브레스": 1}
    if "슈퍼아로마" in text:
        return {"슈퍼아로마": 1}
    if text.startswith("[DH] 담향") or "담향" in text:
        return {"기타": 1}
    return {}


def parse_mm_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    result: dict[str, int] = {}

    if "오일앤크림" in text:
        if any(keyword in text for keyword in ["기프트 세트", "워시", "로션"]):
            add_if(result, "오일앤크림", count_multiple_keywords(text, ["오일앤크림"]))
            if "워시" in text:
                add_if(result, "새싹 워시", count_multiple_keywords(text, ["워시"]))
            if "로션" in text:
                add_if(result, "새싹 로션", count_multiple_keywords(text, ["로션"]))
            if "아토크림" in text:
                add_if(result, "아토크림", count_multiple_keywords(text, ["아토크림"]))
            return result
        return {"오일앤크림": count_pack(text, 1)}

    if "아토크림" in text:
        if any(keyword in text for keyword in ["워시", "로션", "오일앤크림", "수딩세럼"]):
            add_if(result, "아토크림", count_multiple_keywords(text, ["아토크림"]))
            if "워시" in text:
                add_if(result, "새싹 워시", count_multiple_keywords(text, ["워시"]))
            if "로션" in text:
                add_if(result, "새싹 로션", count_multiple_keywords(text, ["로션"]))
            if "오일앤크림" in text:
                add_if(result, "오일앤크림", count_multiple_keywords(text, ["오일앤크림"]))
            if "수딩세럼" in text:
                add_if(result, "기타(젤, 힙, 미니, 세트)", count_multiple_keywords(text, ["수딩세럼"]))
            return result
        return {"아토크림": count_pack(text, 1)}

    if "새싹 로션" in text:
        if any(keyword in text for keyword in ["워시", "오일앤크림", "수딩세럼", "아토크림"]):
            add_if(result, "새싹 로션", count_multiple_keywords(text, ["로션"]))
            if "워시" in text:
                add_if(result, "새싹 워시", count_multiple_keywords(text, ["워시"]))
            if "오일앤크림" in text:
                add_if(result, "오일앤크림", count_multiple_keywords(text, ["오일앤크림"]))
            if "아토크림" in text:
                add_if(result, "아토크림", count_multiple_keywords(text, ["아토크림"]))
            if "수딩세럼" in text:
                add_if(result, "기타(젤, 힙, 미니, 세트)", count_multiple_keywords(text, ["수딩세럼"]))
            return result
        return {"새싹 로션": count_pack(text, 1)}

    if "새싹 워시" in text:
        if any(keyword in text for keyword in ["로션", "오일앤크림", "수딩세럼", "아토크림", "힙 버블 클렌저"]):
            add_if(result, "새싹 워시", count_multiple_keywords(text, ["워시"]))
            if "로션" in text:
                add_if(result, "새싹 로션", count_multiple_keywords(text, ["로션"]))
            if "오일앤크림" in text:
                add_if(result, "오일앤크림", count_multiple_keywords(text, ["오일앤크림"]))
            if "아토크림" in text:
                add_if(result, "아토크림", count_multiple_keywords(text, ["아토크림"]))
            if "수딩세럼" in text or "힙 버블 클렌저" in text:
                add_if(result, "기타(젤, 힙, 미니, 세트)", 1)
            return result
        return {"새싹 워시": count_pack(text, 1)}

    if any(keyword in text for keyword in ["힙 버블 클렌저", "수딩세럼", "미니어처", "트래블키트", "에센셜 키트", "벨리케어 키트", "방수요", "턱받이", "파우치", "가운"]):
        return {"기타(젤, 힙, 미니, 세트)": 1}

    return result


def parse_other_item(name: str) -> dict[str, int]:
    text = clean_name(name)
    if text == "할인" or "포인트 할인" in text:
        return {"할인": 1}
    if "(BT)" in text or "쇼핑백" in text or "진열대" in text or "고객카드" in text or "해면" in text or "힐링 패드" in text or "리프팅컵" in text:
        return {"뷰티툴(BT)": 1}
    if "택배비" in text or "리플렛" in text or "단종" in text or "(CT)" in text or "카탈로그" in text or "밀리카드" in text or "배너" in text or "필름" in text:
        return {"기타(택배비, 리플렛, 단종)": 1}
    return {}


PARSERS: list[Callable[[str], dict[str, int]]] = [
    parse_ta_item,
    parse_ts_item,
    parse_device_item,
    parse_dh_item,
    parse_mm_item,
    parse_other_item,
]


@dataclass
class Totals:
    qty_26: dict[str, int]
    qty_25: dict[str, int]


def aggregate(raw_path: Path) -> Totals:
    df = pd.read_excel(raw_path, sheet_name="판매현황", header=[1, 2])
    name_col = df.columns[0]
    qty_26_col = YEAR26_QTY
    qty_25_col = YEAR25_QTY

    totals_26: dict[str, int] = defaultdict(int)
    totals_25: dict[str, int] = defaultdict(int)

    for _, row in df.iterrows():
        name = row[name_col]
        if pd.isna(name):
            continue

        for parser in PARSERS:
            matched = parser(str(name))
            if not matched:
                continue

            qty_26 = int(row[qty_26_col]) if pd.notna(row[qty_26_col]) else 0
            qty_25 = int(row[qty_25_col]) if pd.notna(row[qty_25_col]) else 0

            for label, multiplier in matched.items():
                totals_26[label] += qty_26 * multiplier
                totals_25[label] += qty_25 * multiplier
            break

    return Totals(dict(totals_26), dict(totals_25))


def write_report(template_path: Path, out_path: Path, totals: Totals, sheet_name: str) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name]

    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 2).value
        if not label or label not in totals.qty_26 and label not in totals.qty_25:
            continue

        qty_26 = totals.qty_26.get(label, 0)
        qty_25 = totals.qty_25.get(label, 0)

        ws.cell(row, 3).value = qty_26 if qty_26 != 0 else None
        ws.cell(row, 4).value = qty_25 if qty_25 != 0 else None

    wb.save(out_path)


def compare_with_template(template_path: Path, totals: Totals, sheet_name: str) -> None:
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb[sheet_name]
    print("\n기존 시트와 차이")
    print("품목\t계산(26)\t기존(26)\t계산(25)\t기존(25)")
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 2).value
        if not label:
            continue
        calc_26 = totals.qty_26.get(label, 0)
        calc_25 = totals.qty_25.get(label, 0)
        cur_26 = ws.cell(row, 3).value
        cur_25 = ws.cell(row, 4).value
        if any(value is not None for value in [cur_26, cur_25]) and (calc_26 != (cur_26 or 0) or calc_25 != (cur_25 or 0)):
            print(f"{label}\t{calc_26}\t{cur_26 or 0}\t{calc_25}\t{cur_25 or 0}")


def print_summary(totals: Totals) -> None:
    keys = sorted(set(totals.qty_26) | set(totals.qty_25))
    print("품목\t2026년 3월\t2025년 3월")
    for key in keys:
        print(f"{key}\t{totals.qty_26.get(key, 0)}\t{totals.qty_25.get(key, 0)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="원본 판매현황 엑셀을 경영회의자료 3월 시트 형식으로 집계합니다.")
    parser.add_argument("--raw", type=Path, required=True, help="26년, 25년 3월 데이터(로우).xlsx 경로")
    parser.add_argument("--template", type=Path, required=True, help="경영회의자료 엑셀 경로")
    parser.add_argument("--out", type=Path, required=True, help="결과 파일 저장 경로")
    parser.add_argument("--sheet", default="3월", help="수정할 시트명")
    args = parser.parse_args()

    totals = aggregate(args.raw)
    write_report(args.template, args.out, totals, args.sheet)
    print_summary(totals)
    compare_with_template(args.template, totals, args.sheet)
    print(f"\n저장 완료: {args.out}")


if __name__ == "__main__":
    main()
