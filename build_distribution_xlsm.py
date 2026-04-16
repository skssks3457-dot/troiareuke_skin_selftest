from __future__ import annotations

from copy import copy
from pathlib import Path
import glob
import os
import shutil

from openpyxl import load_workbook


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
BASE_XLSM = WORKDIR / "sales_hub.xlsm"
OUTPUT_XLSM = WORKDIR / "전사 제품별 판매실적_프로그램.xlsm"

SHEET_HUB = "\ud5c8\ube0c"
SHEET_MASTER = "\uc0c1\ud488\ub9c8\uc2a4\ud130"
SHEET_RULES = "\uc138\ud2b8\uaddc\uce59"
SHEET_MAPPING = "\ubcf4\uace0\uc0c1\ud488\uba85\ucf54\ub4dc\ub9e4\ud551"
SHEET_RAW = "RawData"
SHEET_NORM = "NormalizedData"
SHEET_LOG = "ChangeLog"


def find_source_workbook() -> Path:
    candidates = []
    for p in glob.glob(str(WORKDIR / "*.xlsx")):
        name = os.path.basename(p)
        if name.startswith("~$"):
            continue
        if "\uc804\uc0ac \uc81c\ud488\ubcc4 \ud310\ub9e4\uc2e4\uc801" in name:
            candidates.append(Path(p))
    if not candidates:
        raise FileNotFoundError("Source workbook not found.")
    return max(candidates, key=lambda x: x.stat().st_mtime)


def clear_sheets(wb):
    for idx in range(len(wb.worksheets), 0, -1):
        wb.remove(wb.worksheets[idx - 1])


def copy_sheet_contents(src_ws, dst_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            target = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target.font = copy(cell.font)
                target.fill = copy(cell.fill)
                target.border = copy(cell.border)
                target.alignment = copy(cell.alignment)
                target.number_format = cell.number_format
                target.protection = copy(cell.protection)
            if cell.hyperlink:
                target._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                target.comment = copy(cell.comment)

    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
        dst_ws.column_dimensions[key].hidden = dim.hidden

    for idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[idx].height = dim.height
        dst_ws.row_dimensions[idx].hidden = dim.hidden

    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor


def build_hub(ws):
    ws["A1"] = "\uc804\uc0ac \uc81c\ud488\ubcc4 \ud310\ub9e4\uc2e4\uc801 \ud5c8\ube0c"
    ws["A1"].font = copy(ws["A1"].font)
    ws["A1"].font = ws["A1"].font.copy(bold=True, size=18)
    ws["A3"] = "1. \ub85c\uc6b0 \ub370\uc774\ud130 \ubd88\ub7ec\uc624\uae30"
    ws["A4"] = "- \ub2e4\uc911 \uc120\ud0dd \uac00\ub2a5 (\uc608: 1\uc6d4.xlsx, 2\uc6d4.xlsx, 3\uc6d4.xlsx)"
    ws["A5"] = "- MK004 / \ube0c\ub79c\ub4dc\uc0ac\uc5c5\ubd80\ub9cc \uc138\ud2b8 \ud658\uc0b0"
    ws["A7"] = "2. \uacb0\uacfc \uc0dd\uc131"
    ws["A8"] = "- \ubcf4\uace0\uc0c1\ud488\uba85\ucf54\ub4dc\ub9e4\ud551 \uae30\uc900 \uc9d1\uacc4"
    ws["A9"] = "- \uc2e0\uc0c1\ud488/\uc138\ud2b8/\ub9e4\ud551 \ucd94\uac00 \ud6c4 \ub2e4\uc2dc \uc0dd\uc131"
    ws["A11"] = "3. \uc218\uc815 \uc2dc\ud2b8"
    ws["A12"] = "- \uc0c1\ud488\ub9c8\uc2a4\ud130"
    ws["A13"] = "- \uc138\ud2b8\uaddc\uce59"
    ws["A14"] = "- \ubcf4\uace0\uc0c1\ud488\uba85\ucf54\ub4dc\ub9e4\ud551"
    ws["D3"] = "\ucd5c\uc885 \uae30\uc900\ubcf8"
    ws["E3"] = "\uc804\uc0ac \uc81c\ud488\ubcc4 \ud310\ub9e4\uc2e4\uc801"
    ws["D4"] = "\ub9c8\uc9c0\ub9c9 \ubd88\ub7ec\uc628 \ud30c\uc77c"
    ws["E4"] = "-"
    ws["D5"] = "\ub9c8\uc9c0\ub9c9 \uc2e4\ud589\uc2dc\uac01"
    ws["E5"] = "-"
    ws["D6"] = "\ud604\uc7ac \uc6d4 \uc218"
    ws["E6"] = "-"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 50


def build_helper_sheet(ws, headers):
    ws.append(headers)
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = title


def main():
    source = find_source_workbook()
    shutil.copy2(BASE_XLSM, OUTPUT_XLSM)

    src_wb = load_workbook(source, data_only=False)
    dst_wb = load_workbook(OUTPUT_XLSM, keep_vba=True)
    clear_sheets(dst_wb)

    hub = dst_wb.create_sheet(SHEET_HUB, 0)
    build_hub(hub)

    for name in [SHEET_MASTER, SHEET_RULES, SHEET_MAPPING]:
        new_ws = dst_wb.create_sheet(name)
        copy_sheet_contents(src_wb[name], new_ws)

    raw_ws = dst_wb.create_sheet(SHEET_RAW)
    build_helper_sheet(raw_ws, ["기준월", "연도", "월", "거래처그룹코드", "거래처그룹", "품목코드", "품목명", "수량", "합계"])

    norm_ws = dst_wb.create_sheet(SHEET_NORM)
    build_helper_sheet(norm_ws, ["기준월", "연도", "월", "거래처그룹코드", "거래처그룹", "원본품목코드", "원본품목명", "집계품목코드", "집계품목명", "라인", "보고용상품명", "원본수량", "실물수량", "정산수량", "비교매출", "원본합계", "세트환산여부"])

    log_ws = dst_wb.create_sheet(SHEET_LOG)
    build_helper_sheet(log_ws, ["일시", "사용자", "작업", "대상", "비고"])
    raw_ws.sheet_state = "hidden"
    norm_ws.sheet_state = "hidden"
    log_ws.sheet_state = "hidden"

    dst_wb.save(OUTPUT_XLSM)
    src_wb.close()
    dst_wb.close()


if __name__ == "__main__":
    main()
