from __future__ import annotations

from pathlib import Path
from shutil import copy2
import glob
import os

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
SUMMARY_SHEET = "월간요약"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
HEADER_FILL = PatternFill("solid", fgColor="EAF1FB")
SECTION_FILL = PatternFill("solid", fgColor="F7F9FC")
SUB_FILL = PatternFill("solid", fgColor="F3F6FB")

LINE_TARGETS = [
    "트로이아르케 소계",
    "악센 소계",
    "트로이아르케 전체",
    "서울 소계",
    "미용기기 소계",
    "담향 소계",
    "밀리맘 소계",
    "기타 소계",
]


def latest_program_file() -> Path:
    candidates = [Path(p) for p in glob.glob(str(WORKDIR / "*.xlsm")) if os.path.basename(p) != "sales_hub.xlsm"]
    if not candidates:
        raise FileNotFoundError("program xlsm not found")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def month_labels(wb):
    sheets = [name for name in wb.sheetnames if name.endswith("월") and name[:-1].isdigit()]
    sheets.sort(key=lambda x: int(x[:-1]))
    return sheets


def find_label_row(ws, label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value == label:
            return r
    return None


def pct(curr: float, prev: float) -> float:
    if prev == 0:
        return 0.0 if curr == 0 else 1.0
    return (curr / prev) - 1


def style_range(ws, start_row: int, end_row: int, start_col: int, end_col: int, fill=None, bold=False, center=False):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if bold:
                cell.font = Font(bold=True)
            if center:
                cell.alignment = Alignment(horizontal="center", vertical="center")


def add_bar_chart(ws, title: str, cat_col: int, data_cols: list[int], start_row: int, end_row: int, anchor: str, width: float = 12.5, height: float = 6.5):
    if end_row <= start_row:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.height = height
    chart.width = width
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    for col in data_cols:
        data = Reference(ws, min_col=col, min_row=start_row, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def add_line_chart(ws, title: str, cat_col: int, data_cols: list[int], start_row: int, end_row: int, anchor: str, width: float = 12.5, height: float = 6.5):
    if end_row <= start_row:
        return
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = ""
    chart.height = height
    chart.width = width
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    for col in data_cols:
        data = Reference(ws, min_col=col, min_row=start_row, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def build_monthly_overview(ws, wb, start_row: int) -> int:
    ws.cell(start_row, 1).value = "월별 전체 요약"
    ws.cell(start_row, 1).font = Font(bold=True, size=12)
    start_row += 1

    headers = ["월", "당월수량", "전년수량", "증감", "당월금액", "전년금액", "YOY", "누계수량", "전년누계수량", "누계증감", "누계금액", "전년누계금액", "누계YOY"]
    for idx, text in enumerate(headers, start=1):
        ws.cell(start_row, idx).value = text
    style_range(ws, start_row, start_row, 1, 13, fill=HEADER_FILL, bold=True, center=True)

    row = start_row + 1
    for month_name in month_labels(wb):
        mws = wb[month_name]
        total_row = find_label_row(mws, "전체 소계")
        if not total_row:
            continue
        month_qty = mws.cell(total_row, 3).value or 0
        prev_qty = mws.cell(total_row, 4).value or 0
        month_amt = mws.cell(total_row, 7).value or 0
        prev_amt = mws.cell(total_row, 8).value or 0
        cum_qty = mws.cell(total_row, 11).value or 0
        cum_prev_qty = mws.cell(total_row, 12).value or 0
        cum_amt = mws.cell(total_row, 15).value or 0
        cum_prev_amt = mws.cell(total_row, 16).value or 0
        vals = [
            month_name, month_qty, prev_qty, pct(month_qty, prev_qty),
            month_amt, prev_amt, pct(month_amt, prev_amt),
            cum_qty, cum_prev_qty, pct(cum_qty, cum_prev_qty),
            cum_amt, cum_prev_amt, pct(cum_amt, cum_prev_amt),
        ]
        for col, val in enumerate(vals, start=1):
            ws.cell(row, col).value = val
        row += 1

    style_range(ws, start_row + 1, row - 1, 1, 13)
    for r in range(start_row + 1, row):
        for c in [2, 3, 8, 9]:
            ws.cell(r, c).number_format = "#,##0"
        for c in [5, 6, 11, 12]:
            ws.cell(r, c).number_format = '"₩"#,##0'
        for c in [4, 7, 10, 13]:
            ws.cell(r, c).number_format = "0.0%"

    add_bar_chart(ws, "월별 당월 금액 비교", 1, [5, 6], start_row, row - 1, "O3", width=13.5)
    add_line_chart(ws, "월별 누계 금액 추이", 1, [11, 12], start_row, row - 1, "O19", width=13.5)
    return row + 2


def build_line_visual_sections(ws, wb, start_row: int) -> int:
    for month_name in month_labels(wb):
        mws = wb[month_name]

        ws.cell(start_row, 1).value = f"{month_name} 라인별 요약"
        ws.cell(start_row, 1).font = Font(bold=True, size=12)
        start_row += 1

        headers = ["라인", "당월수량", "전년수량", "증감", "당월금액", "전년금액", "YOY"]
        for idx, text in enumerate(headers, start=1):
            ws.cell(start_row, idx).value = text
        style_range(ws, start_row, start_row, 1, 7, fill=HEADER_FILL, bold=True, center=True)

        data_start = start_row + 1
        row = data_start
        for label in LINE_TARGETS:
            rr = find_label_row(mws, label)
            if not rr:
                continue
            month_qty = mws.cell(rr, 3).value or 0
            prev_qty = mws.cell(rr, 4).value or 0
            month_amt = mws.cell(rr, 7).value or 0
            prev_amt = mws.cell(rr, 8).value or 0
            vals = [label, month_qty, prev_qty, pct(month_qty, prev_qty), month_amt, prev_amt, pct(month_amt, prev_amt)]
            for col, val in enumerate(vals, start=1):
                ws.cell(row, col).value = val
            row += 1

        style_range(ws, data_start, row - 1, 1, 7)
        for r in range(data_start, row):
            for c in [2, 3]:
                ws.cell(r, c).number_format = "#,##0"
            for c in [5, 6]:
                ws.cell(r, c).number_format = '"₩"#,##0'
            for c in [4, 7]:
                ws.cell(r, c).number_format = "0.0%"

        add_bar_chart(ws, f"{month_name} 라인별 당월 금액", 1, [5], start_row, row - 1, f"I{start_row}", width=12.5)
        add_bar_chart(ws, f"{month_name} 라인별 전년 금액", 1, [6], start_row, row - 1, f"Q{start_row}", width=12.5)
        start_row = row + 2

    return start_row


def build_top10_sections(ws, wb, start_row: int) -> int:
    for month_name in month_labels(wb):
        mws = wb[month_name]
        ws.cell(start_row, 1).value = f"{month_name} 상위 품목 TOP10"
        ws.cell(start_row, 1).font = Font(bold=True, size=12)
        start_row += 1

        headers = ["순위", "라인", "보고상품명", "당월수량", "당월금액"]
        for idx, text in enumerate(headers, start=1):
            ws.cell(start_row, idx).value = text
        style_range(ws, start_row, start_row, 1, 5, fill=HEADER_FILL, bold=True, center=True)
        start_row += 1

        items = []
        current_line = ""
        for r in range(5, mws.max_row + 1):
            a = mws.cell(r, 1).value
            b = mws.cell(r, 2).value
            if a:
                current_line = a
            if not b:
                continue
            if b.endswith("소계") or b in ("트로이아르케 전체", "전체 소계"):
                continue
            if b in ("할인", "택배비"):
                continue
            qty = mws.cell(r, 3).value or 0
            amt = mws.cell(r, 7).value or 0
            if qty == 0 and amt == 0:
                continue
            items.append((current_line, b, qty, amt))

        items.sort(key=lambda x: abs(x[3]), reverse=True)
        top_items = items[:10]

        data_start = start_row
        for idx, item in enumerate(top_items, start=1):
            ws.cell(start_row, 1).value = idx
            ws.cell(start_row, 2).value = item[0]
            ws.cell(start_row, 3).value = item[1]
            ws.cell(start_row, 4).value = item[2]
            ws.cell(start_row, 5).value = item[3]
            start_row += 1

        if top_items:
            style_range(ws, data_start, start_row - 1, 1, 5)
            for r in range(data_start, start_row):
                ws.cell(r, 4).number_format = "#,##0"
                ws.cell(r, 5).number_format = '"₩"#,##0'

            add_bar_chart(ws, f"{month_name} TOP10 금액", 3, [5], data_start - 1, start_row - 1, f"H{data_start - 1}", width=12.5)
        start_row += 2

    return start_row


def set_layout(ws):
    widths = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 14,
        "G": 10,
        "H": 3,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 3,
        "O": 14,
        "P": 14,
        "Q": 14,
        "R": 14,
        "S": 14,
        "T": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"
    for r in range(1, ws.max_row + 1):
        for c in range(1, 21):
            ws.cell(r, c).alignment = Alignment(
                vertical="center",
                horizontal="center" if c != 3 else "left",
            )


def main():
    xlsm = latest_program_file()
    backup = xlsm.with_name(xlsm.stem + "_backup_summary.xlsm")
    copy2(xlsm, backup)

    wb = load_workbook(xlsm, keep_vba=True)
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET, 1)

    ws["A1"] = "월간 요약 보고"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:M1")
    style_range(ws, 1, 1, 1, 13, fill=TITLE_FILL)

    next_row = 3
    next_row = build_monthly_overview(ws, wb, next_row)
    next_row = build_line_visual_sections(ws, wb, next_row)
    build_top10_sections(ws, wb, next_row)
    set_layout(ws)

    wb.save(xlsm)
    wb.close()


if __name__ == "__main__":
    main()
