from __future__ import annotations

from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
DESKTOP = Path(r"C:\Users\MK\Desktop")
SOURCE_PATH = WORKDIR / "판매분석_전처리_기초.xlsx"
OUTPUT_PATH = WORKDIR / "경영회의자료_MK_1~3월_자동집계_수정.xlsx"


REPORT_LAYOUT = [
    ("트로이아르케", "트로이필(웨폰키트)"),
    ("트로이아르케", "VVS키트"),
    ("트로이아르케", "힐링칵테일"),
    ("트로이아르케", "악센 리커버리"),
    ("트로이아르케", "악센 오일컷클렌징"),
    ("트로이아르케", "AGT하이드로 크림"),
    ("트로이아르케", "악센 센앰플"),
    ("트로이아르케", "AGT하이드로 에센스"),
    ("트로이아르케", "에너지크림"),
    ("트로이아르케", "쉴드크림"),
    ("트로이아르케", "아이크림"),
    ("트로이아르케", "악센 AC스팟솔루션"),
    ("트로이아르케", "악센 UV 프로텍터에센스"),
    ("트로이아르케", "인텐스 UV프로텍터크림"),
    ("트로이아르케", "멜라C인퓨저"),
    ("트로이아르케", "GPS마스크 3종"),
    ("트로이아르케", "에스테틱 비비크림"),
    ("트로이아르케", "트로이아르케쿠션"),
    ("트로이아르케", "기타(링플, 샤쉐, LD, 악센)"),
    ("TA 서울라인", "엑토인 톤업 선세럼"),
    ("TA 서울라인", "옥시젠 토닝 클렌저"),
    ("TA 서울라인", "워터인 퀀칭 세럼"),
    ("TA 서울라인", "엑토인 퀀칭 크림"),
    ("TA 서울라인", "더마인 리+베리어"),
    ("TA 서울라인", "신부쿠션 21호/22호"),
    ("TA 서울라인", "기타(세트상품)"),
    ("미용기기", "멜라소닉(부속품포함)"),
    ("미용기기", "아크소닉(부속품포함)"),
    ("미용기기", "엘디소닉(부속품포함)"),
    ("미용기기", "AI피부진단기"),
    ("담 향", "소프트브레스"),
    ("담 향", "슈퍼아로마"),
    ("담 향", "기타"),
    ("밀리맘", "오일앤크림"),
    ("밀리맘", "아토크림"),
    ("밀리맘", "새싹 로션"),
    ("밀리맘", "새싹 워시"),
    ("밀리맘", "기타(젤, 힙, 미니, 세트)"),
    ("기 타", "할인"),
    ("기 타", "뷰티툴(BT)"),
    ("기 타", "파트너스(교육)"),
    ("기 타", "기타(택배비, 리플렛, 단종)"),
]

GROUP_ORDER: list[str] = []
for group, _ in REPORT_LAYOUT:
    if group not in GROUP_ORDER:
        GROUP_ORDER.append(group)

ITEM_TO_GROUP = {item: group for group, item in REPORT_LAYOUT}

ALIAS_TO_REPORT = {
    "agt하이드로에센스": "AGT하이드로 에센스",
    "agt하이드로크림": "AGT하이드로 크림",
    "리커버리": "악센 리커버리",
    "오일컷클렌징": "악센 오일컷클렌징",
    "센앰플": "악센 센앰플",
    "ac스팟솔루션": "악센 AC스팟솔루션",
    "gps마스크블루": "GPS마스크 3종",
    "gps마스크레드": "GPS마스크 3종",
    "gps옐로우": "GPS마스크 3종",
    "힐링칵테일": "힐링칵테일",
    "에스테틱비비크림": "에스테틱 비비크림",
    "힐링쿠션": "트로이아르케쿠션",
    "신부쿠션": "신부쿠션 21호/22호",
    "엑토인톤업선세럼": "엑토인 톤업 선세럼",
    "옥시젠토닝클렌저": "옥시젠 토닝 클렌저",
    "워터인퀀칭세럼": "워터인 퀀칭 세럼",
    "엑토인퀀칭크림": "엑토인 퀀칭 크림",
    "더마인리+베리어": "더마인 리+베리어",
    "멜라c인퓨저": "멜라C인퓨저",
    "오일앤크림": "오일앤크림",
    "아토크림": "아토크림",
    "새싹로션": "새싹 로션",
    "새싹워시": "새싹 워시",
    "에너지크림": "에너지크림",
    "쉴드크림": "쉴드크림",
    "아이크림": "아이크림",
    "뷰티툴(bt)": "뷰티툴(BT)",
    "바디med키트": "기타(세트상품)",
    "새싹수딩세럼": "기타(젤, 힙, 미니, 세트)",
    "새싹미니어처": "기타(젤, 힙, 미니, 세트)",
    "에센셜키트": "기타(젤, 힙, 미니, 세트)",
    "기프트세트": "기타(젤, 힙, 미니, 세트)",
    "ac클리어앰플": "기타(링플, 샤쉐, LD, 악센)",
    "ac크림": "기타(링플, 샤쉐, LD, 악센)",
    "셀레믹스세럼": "기타(링플, 샤쉐, LD, 악센)",
    "시카센토너": "기타(링플, 샤쉐, LD, 악센)",
    "포어컨트롤마스크": "기타(링플, 샤쉐, LD, 악센)",
    "휩클렌저": "기타(링플, 샤쉐, LD, 악센)",
    "pit클렌징밀크": "기타(링플, 샤쉐, LD, 악센)",
    "ld오일": "기타(링플, 샤쉐, LD, 악센)",
    "rx앰플세럼": "기타(링플, 샤쉐, LD, 악센)",
    "rx앰플크림": "기타(링플, 샤쉐, LD, 악센)",
    "rx스칼프클렌저": "기타(링플, 샤쉐, LD, 악센)",
    "토닝세럼마스크": "기타(링플, 샤쉐, LD, 악센)",
    "퍼플시카마스크": "기타(링플, 샤쉐, LD, 악센)",
    "페이스핏키트": "기타(링플, 샤쉐, LD, 악센)",
    "선케어": "기타(링플, 샤쉐, LD, 악센)",
    "마일드선크림": "기타(링플, 샤쉐, LD, 악센)",
    "sos슬리핑마스크": "기타(링플, 샤쉐, LD, 악센)",
    "기타": "기타(링플, 샤쉐, LD, 악센)",
}


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return (
        str(value)
        .replace(" ", "")
        .replace("\n", "")
        .replace("\t", "")
        .replace("-", "")
        .replace("_", "")
        .lower()
    )


def safe_num(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def find_compare_files() -> list[tuple[int, Path]]:
    targets = []
    for month in (1, 2, 3):
        path = DESKTOP / f"{month}월 비교 자료.xlsx"
        if not path.exists():
            raise FileNotFoundError(path)
        targets.append((month, path))
    return targets


def load_reference():
    wb = load_workbook(SOURCE_PATH, data_only=True)
    ws_master, ws_rules, ws_map = wb.worksheets[:3]

    master: dict[str, dict] = {}
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        master[str(code).strip()] = {
            "품목명": row[1],
            "라인": row[2],
            "구분": row[3],
            "소비자가": row[4],
            "실물수량기준": safe_num(row[5] or 1),
            "정산수량기준": safe_num(row[6] or 1),
        }

    rules: dict[str, list[dict]] = defaultdict(list)
    for row in ws_rules.iter_rows(min_row=2, values_only=True):
        set_code = row[0]
        comp_code = row[3]
        if not set_code or not comp_code:
            continue
        rules[str(set_code).strip()].append(
            {
                "구성품코드": str(comp_code).strip(),
                "구성품명": row[4],
                "실물수량": safe_num(row[5]),
                "정산수량": safe_num(row[6]),
                "배분비율": safe_num(row[7]),
            }
        )

    code_to_report: dict[str, str] = {}
    for row in ws_map.iter_rows(min_row=2, values_only=True):
        report_name, code = row[0], row[1]
        if report_name and code:
            code_to_report[str(code).strip()] = str(report_name).strip()

    wb.close()
    return master, rules, code_to_report


def parse_month_compare(month: int, path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    records: list[dict] = []

    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 3).value
        name = ws.cell(r, 4).value
        if not code or not name:
            continue

        code = str(code).strip()
        name = str(name).strip()
        qty_2026 = safe_num(ws.cell(r, 5).value)
        amt_2026 = safe_num(ws.cell(r, 8).value)
        qty_2025 = safe_num(ws.cell(r, 9).value)
        amt_2025 = safe_num(ws.cell(r, 12).value)

        if qty_2026 or amt_2026:
            records.append(
                {
                    "기준월": f"2026/{month:02d}",
                    "연도": 2026,
                    "월": month,
                    "품목코드": code,
                    "품목명": name,
                    "수량": qty_2026,
                    "합계": amt_2026,
                }
            )
        if qty_2025 or amt_2025:
            records.append(
                {
                    "기준월": f"2025/{month:02d}",
                    "연도": 2025,
                    "월": month,
                    "품목코드": code,
                    "품목명": name,
                    "수량": qty_2025,
                    "합계": amt_2025,
                }
            )

    wb.close()
    return records


def canonical_report_item(mapped_name: str | None, code: str, name: str, line: str) -> str:
    text = normalize_text(name)
    code_n = normalize_text(code)
    mapped_n = normalize_text(mapped_name)

    if code == "MKTPWPKIT" or "웨폰키트" in name or "weaponkit" in code_n:
        return "트로이필(웨폰키트)"
    if code == "MKRXVVSKIT" or ("vvs" in text and "키트" in name):
        return "VVS키트"
    if "vvs" in text:
        if code.startswith("MKBT"):
            return "뷰티툴(BT)"
        return "기타(링플, 샤쉐, LD, 악센)"
    if "리플렛" in name or "브로슈어" in name:
        return "기타(택배비, 리플렛, 단종)"
    if "택배비" in name or "배송비" in name:
        return "기타(택배비, 리플렛, 단종)"
    if "교육" in name:
        return "파트너스(교육)"
    if code.startswith("MKBT") or "(BT)" in name:
        return "뷰티툴(BT)"

    if mapped_n and mapped_n in ALIAS_TO_REPORT:
        return ALIAS_TO_REPORT[mapped_n]
    if mapped_name in ITEM_TO_GROUP:
        return mapped_name

    if "트로이아르케" in line:
        if "쿠션" in name:
            return "트로이아르케쿠션"
        if "힐링칵테일" in mapped_name if mapped_name else False:
            return "힐링칵테일"
        return "기타(링플, 샤쉐, LD, 악센)"

    if "서울" in line:
        if "쿠션" in name:
            return "신부쿠션 21호/22호"
        return "기타(세트상품)"

    if "밀리맘" in line:
        if "오일앤크림" in name or mapped_n == "오일앤크림":
            return "오일앤크림"
        if "아토크림" in name or mapped_n == "아토크림":
            return "아토크림"
        if "로션" in name or mapped_n == "새싹로션":
            return "새싹 로션"
        if "워시" in name or mapped_n == "새싹워시":
            return "새싹 워시"
        return "기타(젤, 힙, 미니, 세트)"

    return "기타(링플, 샤쉐, LD, 악센)"


def normalize_records(raw_records, master, rules, code_to_report):
    normalized: list[dict] = []
    agg = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})

    for rec in raw_records:
        code = rec["품목코드"]
        base_name = rec["품목명"]
        qty = rec["수량"]
        amt = rec["합계"]

        if code in rules:
            for rule in rules[code]:
                norm_code = rule["구성품코드"]
                master_info = master.get(norm_code, {})
                norm_name = master_info.get("품목명") or rule["구성품명"] or norm_code
                line = master_info.get("라인", "")
                div = master_info.get("구분", "단품")
                mapped_name = code_to_report.get(norm_code)
                report_item = canonical_report_item(mapped_name, norm_code, norm_name, line)
                report_group = ITEM_TO_GROUP[report_item]
                norm_qty = qty * rule["정산수량"]
                norm_actual = qty * rule["실물수량"]
                norm_amt = amt * rule["배분비율"]
                normalized.append(
                    {
                        "기준월": rec["기준월"],
                        "연도": rec["연도"],
                        "월": rec["월"],
                        "원본품목코드": code,
                        "원본품목명": base_name,
                        "정규화품목코드": norm_code,
                        "정규화품목명": norm_name,
                        "라인": line,
                        "구분": div,
                        "보고대분류": report_group,
                        "보고상품명": report_item,
                        "원본수량": qty,
                        "실물수량": norm_actual,
                        "정산수량": norm_qty,
                        "비교매출": norm_amt,
                        "원본합계": amt,
                    }
                )
                agg[(rec["연도"], rec["월"], report_group, report_item)]["qty"] += norm_qty
                agg[(rec["연도"], rec["월"], report_group, report_item)]["amt"] += norm_amt
        else:
            master_info = master.get(code, {})
            line = master_info.get("라인", "")
            div = master_info.get("구분", "단품")
            qty_factor = safe_num(master_info.get("정산수량기준", 1)) or 1
            actual_factor = safe_num(master_info.get("실물수량기준", 1)) or 1
            mapped_name = code_to_report.get(code)
            report_item = canonical_report_item(mapped_name, code, base_name, line)
            report_group = ITEM_TO_GROUP[report_item]
            norm_qty = qty * qty_factor
            norm_actual = qty * actual_factor
            normalized.append(
                {
                    "기준월": rec["기준월"],
                    "연도": rec["연도"],
                    "월": rec["월"],
                    "원본품목코드": code,
                    "원본품목명": base_name,
                    "정규화품목코드": code,
                    "정규화품목명": master_info.get("품목명") or base_name,
                    "라인": line,
                    "구분": div,
                    "보고대분류": report_group,
                    "보고상품명": report_item,
                    "원본수량": qty,
                    "실물수량": norm_actual,
                    "정산수량": norm_qty,
                    "비교매출": amt,
                    "원본합계": amt,
                }
            )
            agg[(rec["연도"], rec["월"], report_group, report_item)]["qty"] += norm_qty
            agg[(rec["연도"], rec["월"], report_group, report_item)]["amt"] += amt

    return normalized, agg


def yoy(current: float, previous: float) -> float | None:
    if previous == 0:
        return None if current == 0 else 1.0
    return (current - previous) / previous


def share(value: float, total: float) -> float:
    if total == 0:
        return 0.0
    return value / total


def add_sheet_copy(wb_out, source_ws, title: str):
    ws_new = wb_out.create_sheet(title)
    for row in source_ws.iter_rows():
        for cell in row:
            ws_new[cell.coordinate].value = cell.value
    for key, dim in source_ws.column_dimensions.items():
        ws_new.column_dimensions[key].width = dim.width
    return ws_new


def write_raw_sheet(ws, rows: list[dict]):
    headers = ["기준월", "연도", "월", "품목코드", "품목명", "수량", "공급가액", "부가세", "합계"]
    ws.append(headers)
    for row in rows:
        ws.append([row["기준월"], row["연도"], row["월"], row["품목코드"], row["품목명"], row["수량"], None, None, row["합계"]])


def write_normalized_sheet(ws, rows: list[dict]):
    headers = [
        "기준월", "연도", "월", "원본품목코드", "원본품목명", "정규화품목코드", "정규화품목명",
        "라인", "구분", "보고대분류", "보고상품명", "원본수량", "실물수량", "정산수량", "비교매출", "원본합계",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])


def build_report_sheet(ws, month: int, agg):
    ws["A1"] = "1. 제품별 판매현황"
    ws["A2"] = "구분"
    ws["C2"] = f"{month}월 당월"
    ws["K2"] = f"{month}월 누계"
    headers = ["수량", "전년", "증감", "구성비", "금액", "전년", "YOY", "구성비", "수량", "전년", "YOY", "구성비", "금액", "전년", "YOY", "구성비"]
    for idx, header in enumerate(headers, start=3):
        ws.cell(3, idx).value = header

    fill = PatternFill("solid", fgColor="F2EDE1")
    for row in range(1, 4):
        for col in range(1, 19):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    totals_by_item = []
    total_month_qty = total_month_amt = total_cum_qty = total_cum_amt = 0.0
    for group, item in REPORT_LAYOUT:
        cur_qty = agg[(2026, month, group, item)]["qty"]
        prev_qty = agg[(2025, month, group, item)]["qty"]
        cur_amt = agg[(2026, month, group, item)]["amt"]
        prev_amt = agg[(2025, month, group, item)]["amt"]
        cur_cum_qty = sum(agg[(2026, m, group, item)]["qty"] for m in range(1, month + 1))
        prev_cum_qty = sum(agg[(2025, m, group, item)]["qty"] for m in range(1, month + 1))
        cur_cum_amt = sum(agg[(2026, m, group, item)]["amt"] for m in range(1, month + 1))
        prev_cum_amt = sum(agg[(2025, m, group, item)]["amt"] for m in range(1, month + 1))
        totals_by_item.append((group, item, cur_qty, prev_qty, cur_amt, prev_amt, cur_cum_qty, prev_cum_qty, cur_cum_amt, prev_cum_amt))
        total_month_qty += cur_qty
        total_month_amt += cur_amt
        total_cum_qty += cur_cum_qty
        total_cum_amt += cur_cum_amt

    row_idx = 4
    current_group = None
    group_rows = []
    all_rows = []
    for group, item, cur_qty, prev_qty, cur_amt, prev_amt, cur_cum_qty, prev_cum_qty, cur_cum_amt, prev_cum_amt in totals_by_item:
        if current_group is None:
            current_group = group
        if group != current_group:
            write_subtotal_row(ws, row_idx, group_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt)
            row_idx += 1
            group_rows = []
            current_group = group
        if not group_rows:
            ws.cell(row_idx, 1).value = group
        ws.cell(row_idx, 2).value = item
        values = [
            cur_qty, prev_qty, yoy(cur_qty, prev_qty), share(cur_qty, total_month_qty),
            cur_amt, prev_amt, yoy(cur_amt, prev_amt), share(cur_amt, total_month_amt),
            cur_cum_qty, prev_cum_qty, yoy(cur_cum_qty, prev_cum_qty), share(cur_cum_qty, total_cum_qty),
            cur_cum_amt, prev_cum_amt, yoy(cur_cum_amt, prev_cum_amt), share(cur_cum_amt, total_cum_amt),
        ]
        for col, value in enumerate(values, start=3):
            ws.cell(row_idx, col).value = value
        group_rows.append(values)
        all_rows.append(values)
        row_idx += 1

    if group_rows:
        write_subtotal_row(ws, row_idx, group_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt)
        row_idx += 1
    write_subtotal_row(ws, row_idx, all_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt, grand=True)
    format_report_sheet(ws, row_idx)


def write_subtotal_row(ws, row_idx, value_rows, total_month_qty, total_month_amt, total_cum_qty, total_cum_amt, grand=False):
    sums = [sum((row[i] or 0) for row in value_rows if isinstance(row[i], (int, float)) or row[i] is None) for i in range(16)]
    values = [
        sums[0], sums[1], yoy(sums[0], sums[1]), share(sums[0], total_month_qty),
        sums[4], sums[5], yoy(sums[4], sums[5]), share(sums[4], total_month_amt),
        sums[8], sums[9], yoy(sums[8], sums[9]), share(sums[8], total_cum_qty),
        sums[12], sums[13], yoy(sums[12], sums[13]), share(sums[12], total_cum_amt),
    ]
    ws.cell(row_idx, 2).value = "합계" if grand else "소계"
    for col, value in enumerate(values, start=3):
        ws.cell(row_idx, col).value = value
    for col in range(1, 19):
        ws.cell(row_idx, col).font = Font(bold=True)


def format_report_sheet(ws, max_row: int):
    ws.freeze_panes = "C4"
    widths = {"A": 18, "B": 28}
    for col in "CDEFGHIJKLMNOPQR":
        widths[col] = 12
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(4, max_row + 1):
        for col in range(3, 19):
            cell = ws.cell(row, col)
            if col in (5, 7, 11, 13, 17, 19):
                pass
            if col in (5, 8, 12, 16):
                cell.number_format = "0.0%"
            elif col in (7, 10, 14, 18):
                cell.number_format = "0.0%"
            elif col in (3, 4, 11, 12):
                cell.number_format = "#,##0"
            elif col in (6, 7, 14, 15):
                cell.number_format = "#,##0"


def main():
    master, rules, code_to_report = load_reference()

    wb_out = Workbook()
    ws = wb_out.active
    wb_out.remove(ws)

    # copy the three reference sheets from source-of-truth workbook
    wb_ref = load_workbook(SOURCE_PATH, data_only=True)
    for sheet in wb_ref.worksheets:
        add_sheet_copy(wb_out, sheet, sheet.title)
    wb_ref.close()

    raw_all = []
    for month, path in find_compare_files():
        rows = parse_month_compare(month, path)
        raw_all.extend(rows)
        ws_raw = wb_out.create_sheet(f"원본_{month}월")
        write_raw_sheet(ws_raw, rows)

    normalized, agg = normalize_records(raw_all, master, rules, code_to_report)
    ws_norm = wb_out.create_sheet("정규화데이터")
    write_normalized_sheet(ws_norm, normalized)

    for month in (1, 2, 3):
        ws_report = wb_out.create_sheet(f"{month}월")
        build_report_sheet(ws_report, month, agg)

    wb_out.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()
