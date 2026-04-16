from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import glob
import os
import re
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKDIR = Path(r"C:\Users\MK\Documents\Playground")
DESKTOP = Path(r"C:\Users\MK\Desktop")
OUTPUT_PATH = WORKDIR / "sales_output.xlsx"

BRAND_GROUP_CODE = "MK004"
BRAND_GROUP_NAME = "\ube0c\ub79c\ub4dc\uc0ac\uc5c5\ubd80"
DEVICE_LINE = "\ubbf8\uc6a9\uae30\uae30"
DEVICE_PARTS_NAME = "\ubbf8\uc6a9\uae30\uae30 \ubd80\uc18d\ud488"
LINE_ORDER = [
    "\ud2b8\ub85c\uc774\uc544\ub974\ucf00",
    "\uc545\uc13c",
    "\uc11c\uc6b8",
    "\ubbf8\uc6a9\uae30\uae30",
    "\ub2f4\ud5a5",
    "\ubc00\ub9ac\ub9d8",
    "\uae30\ud0c0",
]

SHEET_MASTER = "\uc0c1\ud488\ub9c8\uc2a4\ud130"
SHEET_RULES = "\uc138\ud2b8\uaddc\uce59"
SHEET_MAPPING = "\ubcf4\uace0\uc0c1\ud488\uba85\ucf54\ub4dc\ub9e4\ud551"
SHEET_NORMALIZED = "\uc815\uaddc\ud654\ub370\uc774\ud130"
RAW_SHEET_PREFIX = "\uc6d0\ubcf8_"

TITLE_TEXT = "1. \uc81c\ud488\ubcc4 \ud310\ub9e4\ud604\ud669"
LABEL_GROUP = "\uad6c\ubd84"
LABEL_QTY = "\ud310\ub9e4\uc218\ub7c9"
LABEL_AMT = "\ud310\ub9e4\uc2e4\uc801"
LABEL_SUBTOTAL = "\uc18c\uacc4"
LABEL_TROIA_TOTAL = "\ud2b8\ub85c\uc774\uc544\ub974\ucf00 \uc804\uccb4"
LABEL_GRAND_TOTAL = "\uc804\uccb4 \uc18c\uacc4"
LABEL_OTHER = "\uae30\ud0c0"

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
GRAND_FILL = PatternFill("solid", fgColor="C6D9F1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def safe_num(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def pct(curr: float, prev: float) -> float:
    if prev == 0:
        return 0.0 if curr == 0 else 1.0
    return (curr / prev) - 1


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return (
        str(value)
        .replace(" ", "")
        .replace("\n", "")
        .replace("\t", "")
        .replace("-", "")
        .replace("_", "")
        .lower()
    )


def find_source_workbook() -> Path:
    candidates = [Path(p) for p in glob.glob(str(WORKDIR / "*.xlsx")) if "v4" in os.path.basename(p)]
    if not candidates:
        raise FileNotFoundError("No v4 workbook found.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def find_month_files() -> dict[int, Path]:
    out = {}
    pat = re.compile(r"^([123]).*\.xlsx$", re.IGNORECASE)
    for p in glob.glob(str(DESKTOP / "*.xlsx")):
        name = os.path.basename(p)
        m = pat.match(name)
        if not m:
            continue
        month = int(m.group(1))
        pp = Path(p)
        if month not in out or pp.stat().st_mtime > out[month].stat().st_mtime:
            out[month] = pp
    missing = [m for m in (1, 2, 3) if m not in out]
    if missing:
        raise FileNotFoundError(f"Missing month files: {missing}")
    return out


def infer_line(code: str, name: str, fallback: str = "") -> str:
    if fallback:
        return fallback
    code_u = str(code).upper()
    name_s = str(name or "")
    name_n = normalize_text(name)
    if code_u.startswith("MKBD") or "\ubbf8\uc6a9\uae30\uae30" in name_s:
        return DEVICE_LINE
    if code_u.startswith("SDDH") or "\ub2f4\ud5a5" in name_s:
        return "\ub2f4\ud5a5"
    if code_u.startswith("MKMM") or "\ubc00\ub9ac\ub9d8" in name_s:
        return "\ubc00\ub9ac\ub9d8"
    if code_u.startswith("MKAC") or "\uc545\uc13c" in name_s:
        return "\uc545\uc13c"
    if code_u.startswith("MKTS") or "\uc11c\uc6b8" in name_s:
        return "\uc11c\uc6b8"
    if code_u.startswith(("MKRX", "MKTA", "MKTP", "MKBT")) or "\ud2b8\ub85c\uc774\uc544\ub974\ucf00" in name_s or "vvs" in name_n:
        return "\ud2b8\ub85c\uc774\uc544\ub974\ucf00"
    return "\ud2b8\ub85c\uc774\uc544\ub974\ucf00"


def fallback_report_name(line: str) -> str:
    return DEVICE_PARTS_NAME if line == DEVICE_LINE else LABEL_OTHER


def load_reference():
    wb = load_workbook(find_source_workbook(), data_only=True)
    ws_master = wb[SHEET_MASTER]
    ws_rules = wb[SHEET_RULES]
    ws_map = wb[SHEET_MAPPING]

    master = {}
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if not code:
            continue
        master[str(code).strip()] = {
            "name": row[1] or "",
            "line": row[2] or "",
            "div": row[3] or "",
            "price": safe_num(row[4]),
        }

    rules = defaultdict(list)
    for row in ws_rules.iter_rows(min_row=2, values_only=True):
        set_code, _, _, comp_code, comp_name, actual_qty, settle_qty, alloc_ratio = row[:8]
        if not set_code or not comp_code:
            continue
        rules[str(set_code).strip()].append(
            {
                "comp_code": str(comp_code).strip(),
                "comp_name": comp_name or "",
                "actual_qty": safe_num(actual_qty),
                "settle_qty": safe_num(settle_qty),
                "alloc_ratio": safe_num(alloc_ratio),
            }
        )

    code_to_map = {}
    report_order = defaultdict(list)
    seen = defaultdict(set)
    for row in ws_map.iter_rows(min_row=2, values_only=True):
        report_name, code, item_name, line, div, price = row[:6]
        if not code:
            continue
        code = str(code).strip()
        info = {
            "report_name": str(report_name).strip() if report_name else "",
            "item_name": item_name or "",
            "line": str(line).strip() if line else "",
            "div": div or "",
            "price": safe_num(price),
        }
        code_to_map[code] = info
        if info["report_name"] and info["line"] and info["report_name"] not in seen[info["line"]]:
            seen[info["line"]].add(info["report_name"])
            report_order[info["line"]].append(info["report_name"])

    wb.close()
    return master, rules, code_to_map, report_order


def parse_month_file(month: int, path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(4, ws.max_row + 1):
        group_code = ws.cell(r, 1).value
        group_name = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        name = ws.cell(r, 4).value
        if not code or not name:
            continue
        qty_2026 = safe_num(ws.cell(r, 5).value)
        amt_2026 = safe_num(ws.cell(r, 8).value)
        qty_2025 = safe_num(ws.cell(r, 9).value)
        amt_2025 = safe_num(ws.cell(r, 12).value)
        base = {
            "group_code": "" if group_code in (None, "") else str(group_code).strip(),
            "group_name": "" if group_name in (None, "") else str(group_name).strip(),
            "code": str(code).strip(),
            "name": str(name).strip(),
            "month": month,
        }
        if qty_2026 or amt_2026:
            rows.append(base | {"year": 2026, "ym": f"2026/{month:02d}", "qty": qty_2026, "amt": amt_2026})
        if qty_2025 or amt_2025:
            rows.append(base | {"year": 2025, "ym": f"2025/{month:02d}", "qty": qty_2025, "amt": amt_2025})
    wb.close()
    return rows


def normalize_records(raw_records, master, rules, code_to_map):
    normalized = []
    agg = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})

    for rec in raw_records:
        is_brand = rec["group_code"] == BRAND_GROUP_CODE or rec["group_name"] == BRAND_GROUP_NAME
        code = rec["code"]
        name = rec["name"]
        qty = rec["qty"]
        amt = rec["amt"]

        if is_brand and code in rules:
            for rule in rules[code]:
                comp_code = rule["comp_code"]
                master_info = master.get(comp_code, {})
                map_info = code_to_map.get(comp_code, {})
                comp_name = map_info.get("item_name") or master_info.get("name") or rule["comp_name"] or comp_code
                line = map_info.get("line") or infer_line(comp_code, comp_name, master_info.get("line", ""))
                report_name = map_info.get("report_name") or fallback_report_name(line)
                normalized.append(
                    {
                        "ym": rec["ym"],
                        "year": rec["year"],
                        "month": rec["month"],
                        "group_code": rec["group_code"],
                        "group_name": rec["group_name"],
                        "src_code": code,
                        "src_name": name,
                        "item_code": comp_code,
                        "item_name": comp_name,
                        "line": line,
                        "report_name": report_name,
                        "src_qty": qty,
                        "actual_qty": qty * rule["actual_qty"],
                        "settle_qty": qty * rule["settle_qty"],
                        "compare_amt": amt * rule["alloc_ratio"],
                        "src_amt": amt,
                        "expanded": "Y",
                    }
                )
                agg[(rec["year"], rec["month"], line, report_name)]["qty"] += qty * rule["settle_qty"]
                agg[(rec["year"], rec["month"], line, report_name)]["amt"] += amt * rule["alloc_ratio"]
        else:
            map_info = code_to_map.get(code, {})
            master_info = master.get(code, {})
            line = map_info.get("line") or infer_line(code, name, master_info.get("line", ""))
            report_name = map_info.get("report_name") or fallback_report_name(line)
            normalized.append(
                {
                    "ym": rec["ym"],
                    "year": rec["year"],
                    "month": rec["month"],
                    "group_code": rec["group_code"],
                    "group_name": rec["group_name"],
                    "src_code": code,
                    "src_name": name,
                    "item_code": code,
                    "item_name": name,
                    "line": line,
                    "report_name": report_name,
                    "src_qty": qty,
                    "actual_qty": qty,
                    "settle_qty": qty,
                    "compare_amt": amt,
                    "src_amt": amt,
                    "expanded": "N",
                }
            )
            agg[(rec["year"], rec["month"], line, report_name)]["qty"] += qty
            agg[(rec["year"], rec["month"], line, report_name)]["amt"] += amt

    return normalized, agg


def ordered_lines(report_order, agg):
    lines = set(report_order.keys())
    for _, _, line, _ in agg.keys():
        lines.add(line)
    ordered = [line for line in LINE_ORDER if line in lines]
    ordered.extend(sorted(line for line in lines if line not in ordered))
    return ordered


def ordered_report_names(line, report_order, agg):
    names = list(report_order.get(line, []))
    seen = set(names)
    for _, _, agg_line, report_name in agg.keys():
        if agg_line == line and report_name not in seen:
            seen.add(report_name)
            names.append(report_name)
    extra = DEVICE_PARTS_NAME if line == DEVICE_LINE else LABEL_OTHER
    if extra not in seen:
        names.append(extra)
    return names


def month_metrics(agg, line, report_name, month):
    curr = agg.get((2026, month, line, report_name), {"qty": 0.0, "amt": 0.0})
    prev = agg.get((2025, month, line, report_name), {"qty": 0.0, "amt": 0.0})
    curr_cum_qty = curr_cum_amt = prev_cum_qty = prev_cum_amt = 0.0
    for m in range(1, month + 1):
        c = agg.get((2026, m, line, report_name), {"qty": 0.0, "amt": 0.0})
        p = agg.get((2025, m, line, report_name), {"qty": 0.0, "amt": 0.0})
        curr_cum_qty += c["qty"]
        curr_cum_amt += c["amt"]
        prev_cum_qty += p["qty"]
        prev_cum_amt += p["amt"]
    return {
        "month_qty": curr["qty"],
        "month_prev_qty": prev["qty"],
        "month_amt": curr["amt"],
        "month_prev_amt": prev["amt"],
        "cum_qty": curr_cum_qty,
        "cum_prev_qty": prev_cum_qty,
        "cum_amt": curr_cum_amt,
        "cum_prev_amt": prev_cum_amt,
    }


def write_raw_sheet(wb, title, records):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ["기준월", "연도", "월", "거래처그룹코드", "거래처그룹", "품목코드", "품목명", "수량", "합계"]
    ws.append(headers)
    for rec in records:
        ws.append([rec["ym"], rec["year"], rec["month"], rec["group_code"], rec["group_name"], rec["code"], rec["name"], rec["qty"], rec["amt"]])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_normalized_sheet(wb, normalized):
    if SHEET_NORMALIZED in wb.sheetnames:
        del wb[SHEET_NORMALIZED]
    ws = wb.create_sheet(SHEET_NORMALIZED)
    headers = [
        "기준월", "연도", "월", "거래처그룹코드", "거래처그룹",
        "원본품목코드", "원본품목명", "집계품목코드", "집계품목명",
        "라인", "보고용상품명", "원본수량", "실물수량", "정산수량", "비교매출", "원본합계", "세트환산여부",
    ]
    ws.append(headers)
    for rec in normalized:
        ws.append([
            rec["ym"], rec["year"], rec["month"], rec["group_code"], rec["group_name"],
            rec["src_code"], rec["src_name"], rec["item_code"], rec["item_name"],
            rec["line"], rec["report_name"], rec["src_qty"], rec["actual_qty"], rec["settle_qty"],
            rec["compare_amt"], rec["src_amt"], rec["expanded"],
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def fill_headers(ws, month):
    ws.merge_cells("A1:R1")
    ws["A1"] = TITLE_TEXT
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:B4")
    ws["A2"] = LABEL_GROUP
    ws.merge_cells("C2:J2")
    ws["C2"] = f"{month}\uc6d4 \ub2f9\uc6d4"
    ws.merge_cells("K2:R2")
    ws["K2"] = f"{month}\uc6d4 \ub204\uacc4"
    ws.merge_cells("C3:F3")
    ws["C3"] = LABEL_QTY
    ws.merge_cells("G3:J3")
    ws["G3"] = LABEL_AMT
    ws.merge_cells("K3:N3")
    ws["K3"] = LABEL_QTY
    ws.merge_cells("O3:R3")
    ws["O3"] = LABEL_AMT

    labels = [
        "\uc218\ub7c9", "\uc804\ub144", "\uc99d\uac10", "\uad6c\uc131\ube44",
        "\uae08\uc561", "\uc804\ub144", "YOY", "\uad6c\uc131\ube44",
        "\uc218\ub7c9", "\uc804\ub144", "YOY", "\uad6c\uc131\ube44",
        "\uae08\uc561", "\uc804\ub144", "YOY", "\uad6c\uc131\ube44",
    ]
    for idx, label in enumerate(labels, start=3):
        ws.cell(4, idx).value = label

    for row in ws["A2:R4"]:
        for cell in row:
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 14, "B": 22, "C": 9, "D": 9, "E": 9, "F": 9,
        "G": 16, "H": 16, "I": 9, "J": 9, "K": 9, "L": 9,
        "M": 9, "N": 9, "O": 16, "P": 16, "Q": 9, "R": 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "C5"


def write_value_row(ws, row_idx, line_label, report_name, metrics, totals, fill, bold=False):
    qty_share = 0 if totals["month_qty"] == 0 else metrics["month_qty"] / totals["month_qty"]
    amt_share = 0 if totals["month_amt"] == 0 else metrics["month_amt"] / totals["month_amt"]
    cum_qty_share = 0 if totals["cum_qty"] == 0 else metrics["cum_qty"] / totals["cum_qty"]
    cum_amt_share = 0 if totals["cum_amt"] == 0 else metrics["cum_amt"] / totals["cum_amt"]
    values = [
        line_label, report_name,
        metrics["month_qty"], metrics["month_prev_qty"], pct(metrics["month_qty"], metrics["month_prev_qty"]), qty_share,
        metrics["month_amt"], metrics["month_prev_amt"], pct(metrics["month_amt"], metrics["month_prev_amt"]), amt_share,
        metrics["cum_qty"], metrics["cum_prev_qty"], pct(metrics["cum_qty"], metrics["cum_prev_qty"]), cum_qty_share,
        metrics["cum_amt"], metrics["cum_prev_amt"], pct(metrics["cum_amt"], metrics["cum_prev_amt"]), cum_amt_share,
    ]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx >= 3 else "left", vertical="center")
        if bold:
            cell.font = Font(bold=True)
    for col_idx in (3, 4, 7, 8, 11, 12, 15, 16):
        ws.cell(row_idx, col_idx).number_format = '#,##0'
    for col_idx in (5, 6, 9, 10, 13, 14, 17, 18):
        ws.cell(row_idx, col_idx).number_format = '0.0%'
    for col_idx in (7, 8, 15, 16):
        ws.cell(row_idx, col_idx).number_format = '"₩"#,##0'


def merge_line_labels(ws, ranges):
    for line_name, start_row, end_row in ranges:
        if start_row > end_row:
            continue
        if start_row < end_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        cell = ws.cell(start_row, 1)
        cell.value = line_name
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in range(start_row, end_row + 1):
            ws.cell(row, 1).border = BORDER


def build_month_sheet(wb, month, agg, report_order):
    title = f"{month}\uc6d4"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    fill_headers(ws, month)

    lines = ordered_lines(report_order, agg)
    all_metrics = []
    for line in lines:
        for report_name in ordered_report_names(line, report_order, agg):
            all_metrics.append(month_metrics(agg, line, report_name, month))

    totals = {
        "month_qty": sum(x["month_qty"] for x in all_metrics),
        "month_amt": sum(x["month_amt"] for x in all_metrics),
        "cum_qty": sum(x["cum_qty"] for x in all_metrics),
        "cum_amt": sum(x["cum_amt"] for x in all_metrics),
    }

    row_idx = 5
    line_ranges = []
    troia_bucket = []
    troia_lines = {"\ud2b8\ub85c\uc774\uc544\ub974\ucf00", "\uc545\uc13c"}

    for line in lines:
        start_row = row_idx
        line_metrics = []
        for report_name in ordered_report_names(line, report_order, agg):
            metrics = month_metrics(agg, line, report_name, month)
            write_value_row(ws, row_idx, "", report_name, metrics, totals, WHITE_FILL)
            line_metrics.append(metrics)
            if line in troia_lines:
                troia_bucket.append(metrics)
            row_idx += 1
        end_row = row_idx - 1
        line_ranges.append((line, start_row, end_row))

        subtotal = {
            "month_qty": sum(x["month_qty"] for x in line_metrics),
            "month_prev_qty": sum(x["month_prev_qty"] for x in line_metrics),
            "month_amt": sum(x["month_amt"] for x in line_metrics),
            "month_prev_amt": sum(x["month_prev_amt"] for x in line_metrics),
            "cum_qty": sum(x["cum_qty"] for x in line_metrics),
            "cum_prev_qty": sum(x["cum_prev_qty"] for x in line_metrics),
            "cum_amt": sum(x["cum_amt"] for x in line_metrics),
            "cum_prev_amt": sum(x["cum_prev_amt"] for x in line_metrics),
        }
        write_value_row(ws, row_idx, "", f"{line} {LABEL_SUBTOTAL}", subtotal, totals, SUBTOTAL_FILL, bold=True)
        row_idx += 1

        if line == "\uc545\uc13c":
            troia_total = {
                "month_qty": sum(x["month_qty"] for x in troia_bucket),
                "month_prev_qty": sum(x["month_prev_qty"] for x in troia_bucket),
                "month_amt": sum(x["month_amt"] for x in troia_bucket),
                "month_prev_amt": sum(x["month_prev_amt"] for x in troia_bucket),
                "cum_qty": sum(x["cum_qty"] for x in troia_bucket),
                "cum_prev_qty": sum(x["cum_prev_qty"] for x in troia_bucket),
                "cum_amt": sum(x["cum_amt"] for x in troia_bucket),
                "cum_prev_amt": sum(x["cum_prev_amt"] for x in troia_bucket),
            }
            write_value_row(ws, row_idx, "", LABEL_TROIA_TOTAL, troia_total, totals, GRAND_FILL, bold=True)
            row_idx += 1

    overall = {
        "month_qty": sum(x["month_qty"] for x in all_metrics),
        "month_prev_qty": sum(x["month_prev_qty"] for x in all_metrics),
        "month_amt": sum(x["month_amt"] for x in all_metrics),
        "month_prev_amt": sum(x["month_prev_amt"] for x in all_metrics),
        "cum_qty": sum(x["cum_qty"] for x in all_metrics),
        "cum_prev_qty": sum(x["cum_prev_qty"] for x in all_metrics),
        "cum_amt": sum(x["cum_amt"] for x in all_metrics),
        "cum_prev_amt": sum(x["cum_prev_amt"] for x in all_metrics),
    }
    write_value_row(ws, row_idx, "", LABEL_GRAND_TOTAL, overall, totals, GRAND_FILL, bold=True)
    merge_line_labels(ws, line_ranges)


def main():
    source_path = find_source_workbook()
    month_files = find_month_files()
    master, rules, code_to_map, report_order = load_reference()

    raw_by_month = {}
    raw_all = []
    for month, path in month_files.items():
        recs = parse_month_file(month, path)
        raw_by_month[month] = recs
        raw_all.extend(recs)

    normalized, agg = normalize_records(raw_all, master, rules, code_to_map)

    shutil.copy2(source_path, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    for title in [f"{RAW_SHEET_PREFIX}1\uc6d4", f"{RAW_SHEET_PREFIX}2\uc6d4", f"{RAW_SHEET_PREFIX}3\uc6d4", SHEET_NORMALIZED, "1\uc6d4", "2\uc6d4", "3\uc6d4"]:
        if title in wb.sheetnames:
            del wb[title]

    write_raw_sheet(wb, f"{RAW_SHEET_PREFIX}1\uc6d4", raw_by_month[1])
    write_raw_sheet(wb, f"{RAW_SHEET_PREFIX}2\uc6d4", raw_by_month[2])
    write_raw_sheet(wb, f"{RAW_SHEET_PREFIX}3\uc6d4", raw_by_month[3])
    write_normalized_sheet(wb, normalized)
    build_month_sheet(wb, 1, agg, report_order)
    build_month_sheet(wb, 2, agg, report_order)
    build_month_sheet(wb, 3, agg, report_order)
    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()
