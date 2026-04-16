from __future__ import annotations

import re
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_PATH = Path(r"C:\Users\MK\Documents\Playground\판매분석_전처리_기초.xlsx")
RAW_PRODUCT_PATH = Path(r"C:\Users\MK\Desktop\s7mqmLVV4pmBXzhH.xlsx")
VALIDATION_PATH = Path(r"C:\Users\MK\Desktop\qiDJlwun3rUVrKwG.xlsx")
OUTPUT_PATH = Path(r"C:\Users\MK\Documents\Playground\판매분석_월별보고_결과.xlsx")

RETAIL_DEPTS = {"트로이아르케사업팀", "셀럽B2B", "밀리맘사업팀"}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")

NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


@dataclass
class MasterRow:
    code: str
    name: str
    line: str
    division: str
    division_reason: str
    avg_price: float | None
    retail_price: float | None
    physical_qty: float
    settle_qty: float
    status: str
    note: str


@dataclass
class RuleRow:
    set_code: str
    set_name: str
    line: str
    rule_type: str
    component_code: str
    component_name: str
    physical_qty: float
    settle_qty: float
    allocation_ratio: float
    status: str
    note: str


@dataclass
class RawRecord:
    month_key: str
    year: int
    month: int
    dept_code: str
    dept_name: str
    code: str
    name: str
    qty: float
    supply: float
    vat: float
    total: float


@dataclass
class NormalizedRecord:
    month_key: str
    year: int
    month: int
    dept_name: str
    source_code: str
    source_name: str
    target_code: str
    target_name: str
    line: str
    division: str
    report_group: str
    report_item: str
    is_other: str
    source_qty: float
    physical_qty: float
    settle_qty: float
    retail_price: float | None
    compare_amount: float
    raw_total: float
    calc_method: str
    note: str


REPORT_STRUCTURE = [
    ("트로이아르케", ["트로이필(웨폰키트)", "VVS키트", "힐링칵테일", "악센 리커버리", "악센 오일컷클렌징", "AGT하이드로 크림", "악센 센앰플", "AGT하이드로 에센스", "에너지크림", "쉴드크림", "아이크림", "악센 AC스팟솔루션", "악센 UV 프로텍터에센스", "인텐스 UV프로텍터크림", "멜라C인퓨저", "GPS마스크 3종", "에스테틱 비비크림", "트로이아르케쿠션", "기타(링플, 샤쉐, LD, 악센)"]),
    ("TA 서울라인", ["엑토인 톤업 선세럼", "옥시젠 토닝 클렌저", "워터인 퀀칭 세럼", "엑토인 퀀칭 크림", "더마인 리+베리어", "신부쿠션 21호/22호", "기타(세트상품)"]),
    ("미용기기", ["멜라소닉(부속품포함)", "아크소닉(부속품포함)", "엘디소닉(부속품포함)", "AI피부진단기"]),
    ("담 향", ["소프트브레스", "슈퍼아로마", "기타"]),
    ("밀리맘", ["오일앤크림", "아토크림", "새싹 로션", "새싹 워시", "기타(젤, 힙, 미니, 세트)"]),
    ("기 타", ["할인", "뷰티툴(BT)", "파트너스(교육)", "기타(택배비, 리플렛, 단종)"]),
]


def as_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    return float(text)


def style_header(ws, row_no: int = 1) -> None:
    for cell in ws[row_no]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def auto_width(ws) -> None:
    for column in ws.columns:
        letter = column[0].column_letter
        width = 10
        for cell in column:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, len(text) + 2)
        ws.column_dimensions[letter].width = min(width, 36)


def safe_ratio(current: float, previous: float) -> float | None:
    if previous == 0:
        return None
    return (current - previous) / previous


def safe_share(value: float, total: float) -> float | None:
    if total == 0:
        return None
    return value / total


def copy_sheet_values(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(list(row))


def normalize_text(text: str) -> str:
    value = text.upper()
    value = re.sub(r"\([^)]*\)", " ", value)
    value = re.sub(r"\[[^\]]*\]", " ", value)
    value = re.sub(r"\d+\s*ML|\d+\s*G|\d+\s*EA|\d+\s*개|\d+\s*매", " ", value, flags=re.IGNORECASE)
    value = re.sub(r"[-+/_,.*]", " ", value)
    value = re.sub(r"\s+", "", value)
    return value


def strip_variant_suffix(code: str) -> list[str]:
    candidates = [code]
    current = code
    while True:
        nxt = re.sub(r"([-/][A-Z0-9+]+)$", "", current)
        if nxt == current or not nxt:
            break
        candidates.append(nxt)
        current = nxt
    if "-" in code:
        candidates.append(code.split("-")[0])
    if "/" in code:
        candidates.append(code.split("/")[0])
    return list(dict.fromkeys(candidates))


def infer_line(code: str, name: str, master_line: str = "") -> str:
    if master_line:
        return master_line
    upper = name.upper()
    if code.startswith("MKMM") or "밀리맘" in name:
        return "밀리맘"
    if "담향" in name or "소프트브레스" in name or "슈퍼아로마" in name:
        return "담향"
    if code.startswith("MKBD") or "미용기기" in name or "피부진단" in name or "소닉" in name:
        return "미용기기"
    if code.startswith("MKTS") or "서울" in name:
        return "서울"
    if code.startswith("MKBT") or code.startswith("BT") or "(BT)" in name:
        return "뷰티툴"
    if code.startswith("MKAC") or code.startswith("AC") or "악센" in name:
        return "악센"
    if "AI" in upper or "피부진단" in name:
        return "미용기기"
    return "트로이아르케"


def parse_sheet_xml(path: Path) -> list[dict[str, str]]:
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
    rows = []
    for row in root.find("a:sheetData", NS).findall("a:row", NS):
        values: dict[str, str] = {}
        for cell in row.findall("a:c", NS):
            ref = cell.attrib.get("r", "")
            col = "".join(ch for ch in ref if ch.isalpha())
            cell_type = cell.attrib.get("t")
            if cell_type == "inlineStr":
                node = cell.find("a:is/a:t", NS)
                value = node.text if node is not None else ""
            else:
                node = cell.find("a:v", NS)
                value = node.text if node is not None else ""
            values[col] = value
        rows.append(values)
    return rows


def load_master_rows(path: Path) -> dict[str, MasterRow]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["상품마스터"]
    result: dict[str, MasterRow] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0] or "").strip()
        if not code:
            continue
        result[code] = MasterRow(
            code=code,
            name=str(row[1] or "").strip(),
            line=str(row[2] or "").strip(),
            division=str(row[3] or "").strip(),
            division_reason=str(row[4] or "").strip(),
            avg_price=as_float(row[5]) if row[5] not in (None, "") else None,
            retail_price=as_float(row[6]) if row[6] not in (None, "") else None,
            physical_qty=as_float(row[7]) or 1.0,
            settle_qty=as_float(row[8]) or 1.0,
            status=str(row[9] or "").strip(),
            note=str(row[10] or "").strip(),
        )
    return result


def load_rule_rows(path: Path) -> dict[str, list[RuleRow]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["세트규칙"]
    result: dict[str, list[RuleRow]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        set_code = str(row[0] or "").strip()
        if not set_code:
            continue
        result[set_code].append(
            RuleRow(
                set_code=set_code,
                set_name=str(row[1] or "").strip(),
                line=str(row[2] or "").strip(),
                rule_type=str(row[3] or "").strip(),
                component_code=str(row[4] or "").strip(),
                component_name=str(row[5] or "").strip(),
                physical_qty=as_float(row[6]) or 1.0,
                settle_qty=as_float(row[7]) or 1.0,
                allocation_ratio=as_float(row[8]) or 0.0,
                status=str(row[9] or "").strip(),
                note=str(row[10] or "").strip(),
            )
        )
    return result


def parse_product_records(path: Path) -> list[RawRecord]:
    rows = parse_sheet_xml(path)
    records: list[RawRecord] = []
    for row in rows[3:]:
        month_key = (row.get("A") or "").strip()
        if not re.match(r"^\d{4}/\d{2}$", month_key):
            continue
        code = (row.get("D") or "").strip()
        name = (row.get("E") or "").strip()
        if not code:
            continue
        year = int(month_key[:4])
        month = int(month_key[-2:])
        qty_col, supply_col, vat_col, total_col = ("F", "G", "H", "I") if year == 2026 else ("J", "K", "L", "M")
        qty = as_float(row.get(qty_col))
        supply = as_float(row.get(supply_col))
        vat = as_float(row.get(vat_col))
        total = as_float(row.get(total_col))
        if qty == 0 and supply == 0 and vat == 0 and total == 0:
            continue
        records.append(
            RawRecord(
                month_key=month_key,
                year=year,
                month=month,
                dept_code=(row.get("B") or "").strip(),
                dept_name=(row.get("C") or "").strip(),
                code=code,
                name=name,
                qty=qty,
                supply=supply,
                vat=vat,
                total=total,
            )
        )
    return records


def parse_validation_records(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    rows = parse_sheet_xml(path)
    records: list[dict[str, object]] = []
    for row in rows[3:]:
        month_key = (row.get("A") or "").strip()
        dept_name = (row.get("C") or "").strip()
        if not re.match(r"^\d{4}/\d{2}$", month_key) or not dept_name:
            continue
        year = int(month_key[:4])
        qty_col, total_col = ("F", "I") if year == 2026 else ("J", "M")
        qty = as_float(row.get(qty_col))
        total = as_float(row.get(total_col))
        if qty == 0 and total == 0:
            continue
        records.append({"month_key": month_key, "year": year, "month": int(month_key[-2:]), "dept_name": dept_name, "qty": qty, "total": total})
    return records


def build_master_lookup(master_rows: dict[str, MasterRow]) -> dict[str, MasterRow]:
    lookup = dict(master_rows)
    for code, row in master_rows.items():
        for candidate in strip_variant_suffix(code):
            lookup.setdefault(candidate, row)
    return lookup


def map_report_bucket(line: str, code: str, name: str, dept_name: str) -> tuple[str, str, str, str]:
    normalized = normalize_text(name)
    upper = name.upper()

    if "포인트할인" in normalized or code == "GT0002" or name.strip() == "할인":
        return "기 타", "할인", "", "할인 항목"
    if "택배비" in name or "리플렛" in name or "브로슈어" in name or "단종" in name or "쇼핑백" in name or "택배박스" in name:
        return "기 타", "기타(택배비, 리플렛, 단종)", "Y", "기타 비용성 품목"
    if "교육" in dept_name:
        return "기 타", "파트너스(교육)", "", "교육영업팀 귀속"
    if line == "뷰티툴":
        return "기 타", "뷰티툴(BT)", "", "뷰티툴 귀속"
    if line == "미용기기" or "미용기기" in name or "소닉" in name or "피부진단" in name:
        if "멜라소닉" in name:
            return "미용기기", "멜라소닉(부속품포함)", "", "기기명 기준"
        if "아크소닉" in name:
            return "미용기기", "아크소닉(부속품포함)", "", "기기명 기준"
        if "엘디소닉" in name or "LD" in upper:
            return "미용기기", "엘디소닉(부속품포함)", "", "기기명 기준"
        if "피부진단" in name or "AI" in upper:
            return "미용기기", "AI피부진단기", "", "기기명 기준"
        return "미용기기", "멜라소닉(부속품포함)", "Y", "미분류 기기"
    if line == "담향" or "소프트브레스" in name or "슈퍼아로마" in name or "담향" in name:
        if "소프트브레스" in name:
            return "담 향", "소프트브레스", "", "상품명 기준"
        if "슈퍼아로마" in name:
            return "담 향", "슈퍼아로마", "", "상품명 기준"
        return "담 향", "기타", "Y", "담향 기타"
    if line == "밀리맘":
        if "오일앤크림" in name:
            return "밀리맘", "오일앤크림", "", "상품명 기준"
        if "아토크림" in name:
            return "밀리맘", "아토크림", "", "상품명 기준"
        if "로션" in name:
            return "밀리맘", "새싹 로션", "", "상품명 기준"
        if "워시" in name:
            return "밀리맘", "새싹 워시", "", "상품명 기준"
        return "밀리맘", "기타(젤, 힙, 미니, 세트)", "Y", "밀리맘 기타"
    if line == "서울":
        if "톤업 선세럼" in name:
            return "TA 서울라인", "엑토인 톤업 선세럼", "", "상품명 기준"
        if "옥시젠 토닝 클렌저" in name:
            return "TA 서울라인", "옥시젠 토닝 클렌저", "", "상품명 기준"
        if "워터인 퀀칭 세럼" in name:
            return "TA 서울라인", "워터인 퀀칭 세럼", "", "상품명 기준"
        if "엑토인 퀀칭 크림" in name:
            return "TA 서울라인", "엑토인 퀀칭 크림", "", "상품명 기준"
        if "더마인 리" in name or "베리어" in name:
            return "TA 서울라인", "더마인 리+베리어", "", "상품명 기준"
        if "쿠션" in name:
            return "TA 서울라인", "신부쿠션 21호/22호", "", "상품명 기준"
        return "TA 서울라인", "기타(세트상품)", "Y", "서울 기타/세트"

    if "웨폰키트" in name or "트로이필" in name:
        return "트로이아르케", "트로이필(웨폰키트)", "", "상품명 기준"
    if "VVS" in upper:
        return "트로이아르케", "VVS키트", "", "상품명 기준"
    if "힐링칵테일" in name or "H+ 칵테일" in name or "힐링 캡슐" in name:
        return "트로이아르케", "힐링칵테일", "", "상품명 기준"
    if "리커버리" in name:
        return "트로이아르케", "악센 리커버리", "", "상품명 기준"
    if "오일컷클렌징" in name:
        return "트로이아르케", "악센 오일컷클렌징", "", "상품명 기준"
    if "AGT" in upper and "크림" in name:
        return "트로이아르케", "AGT하이드로 크림", "", "상품명 기준"
    if "센앰플" in normalized or "센 앰플" in name:
        return "트로이아르케", "악센 센앰플", "", "상품명 기준"
    if "AGT" in upper and ("에센스" in name or "ESSENCE" in upper):
        return "트로이아르케", "AGT하이드로 에센스", "", "상품명 기준"
    if "에너지크림" in name:
        return "트로이아르케", "에너지크림", "", "상품명 기준"
    if "쉴드크림" in name:
        return "트로이아르케", "쉴드크림", "", "상품명 기준"
    if "아이크림" in name or "아이패치" in name:
        return "트로이아르케", "아이크림", "", "상품명 기준"
    if "AC스팟" in name or "ACSPOT" in upper:
        return "트로이아르케", "악센 AC스팟솔루션", "", "상품명 기준"
    if "UV프로텍터에센스" in normalized or "UV 프로텍터에센스" in name:
        return "트로이아르케", "악센 UV 프로텍터에센스", "", "상품명 기준"
    if "인텐스" in name and "UV프로텍터" in normalized:
        return "트로이아르케", "인텐스 UV프로텍터크림", "", "상품명 기준"
    if "멜라" in name and "인퓨저" in name:
        return "트로이아르케", "멜라C인퓨저", "", "상품명 기준"
    if "GPS" in upper:
        return "트로이아르케", "GPS마스크 3종", "", "상품명 기준"
    if "비비크림" in name:
        return "트로이아르케", "에스테틱 비비크림", "", "상품명 기준"
    if "쿠션" in name:
        return "트로이아르케", "트로이아르케쿠션", "", "상품명 기준"
    return "트로이아르케", "기타(링플, 샤쉐, LD, 악센)", "Y", "트로이아르케 기타"


def expand_record(record: RawRecord, master_lookup: dict[str, MasterRow], rules_by_set: dict[str, list[RuleRow]]) -> list[NormalizedRecord]:
    if not record.dept_name or record.dept_name == "무상출고" or "무상출고" in record.dept_name or record.dept_name.endswith("계"):
        return []

    base_row = None
    for candidate in strip_variant_suffix(record.code):
        if candidate in master_lookup:
            base_row = master_lookup[candidate]
            break

    default_line = infer_line(record.code, record.name, base_row.line if base_row else "")
    default_division = base_row.division if base_row else "단품"

    if record.dept_name not in RETAIL_DEPTS:
        group, item, is_other, map_note = map_report_bucket(default_line, record.code, record.name, record.dept_name)
        return [NormalizedRecord(record.month_key, record.year, record.month, record.dept_name, record.source_code if hasattr(record, 'source_code') else record.code, record.name, record.code, record.name, default_line, default_division, group, item, is_other, record.qty, record.qty, record.qty, base_row.retail_price if base_row else None, record.total, record.total, "원본금액 사용", map_note)]

    rules = [rule for rule in rules_by_set.get(record.code, []) if rule.component_code]
    if rules:
        expanded: list[NormalizedRecord] = []
        for rule in rules:
            component_master = master_lookup.get(rule.component_code)
            line = infer_line(rule.component_code, rule.component_name or record.name, component_master.line if component_master else default_line)
            group, item, is_other, map_note = map_report_bucket(line, rule.component_code, rule.component_name or record.name, record.dept_name)
            retail_price = component_master.retail_price if component_master else None
            settle_qty = record.qty * (rule.settle_qty or 1.0)
            compare_amount = settle_qty * retail_price if retail_price else record.total * (rule.allocation_ratio or 0.0)
            method = "정산수량 x 소비자가" if retail_price else "배분비율 x 원본금액"
            expanded.append(NormalizedRecord(record.month_key, record.year, record.month, record.dept_name, record.code, record.name, rule.component_code, rule.component_name or (component_master.name if component_master else record.name), line, "세트", group, item, is_other, record.qty, record.qty * (rule.physical_qty or 1.0), settle_qty, retail_price, compare_amount, record.total, method, f"{rule.rule_type} / {map_note}"))
        return expanded

    retail_price = base_row.retail_price if base_row else None
    settle_factor = base_row.settle_qty if base_row else 1.0
    physical_factor = base_row.physical_qty if base_row else 1.0
    settle_qty = record.qty * settle_factor
    compare_amount = settle_qty * retail_price if retail_price else record.total
    method = "정산수량 x 소비자가" if retail_price else "원본금액 대체"
    group, item, is_other, map_note = map_report_bucket(default_line, record.code, record.name, record.dept_name)
    return [NormalizedRecord(record.month_key, record.year, record.month, record.dept_name, record.code, record.name, base_row.code if base_row else record.code, base_row.name if base_row else record.name, default_line, default_division, group, item, is_other, record.qty, record.qty * physical_factor, settle_qty, retail_price, compare_amount, record.total, method, map_note)]


def aggregate_validation(records: list[NormalizedRecord]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "amount": 0.0})
    for record in records:
        key = (record.month_key, record.dept_name)
        grouped[key]["qty"] += record.settle_qty if record.dept_name in RETAIL_DEPTS else record.source_qty
        grouped[key]["amount"] += record.compare_amount
    return [{"month_key": month_key, "dept_name": dept_name, "qty": vals["qty"], "total": vals["amount"]} for (month_key, dept_name), vals in sorted(grouped.items())]


def build_mapping_rows(records: list[NormalizedRecord]) -> list[list[object]]:
    seen: dict[str, list[object]] = {}
    for record in records:
        if record.target_code not in seen:
            seen[record.target_code] = [record.target_code, record.target_name, record.line, record.report_group, record.report_item, record.is_other, record.note]
    return list(seen.values())


def build_report_aggregates(records: list[NormalizedRecord]) -> dict[tuple[int, int, str, str], dict[str, float]]:
    grouped: dict[tuple[int, int, str, str], dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "amount": 0.0})
    for record in records:
        grouped[(record.year, record.month, record.report_group, record.report_item)]["qty"] += record.settle_qty if record.dept_name in RETAIL_DEPTS else record.source_qty
        grouped[(record.year, record.month, record.report_group, record.report_item)]["amount"] += record.compare_amount
    return grouped


def extract_report_rows(grouped: dict[tuple[int, int, str, str], dict[str, float]], year: int, month: int) -> tuple[dict[tuple[str, str], dict[str, float]], dict[tuple[str, str], dict[str, float]], dict[str, float], dict[str, float]]:
    current_rows: dict[tuple[str, str], dict[str, float]] = {}
    cumulative_rows: dict[tuple[str, str], dict[str, float]] = {}
    total_current = {"qty": 0.0, "amount": 0.0}
    total_cumulative = {"qty": 0.0, "amount": 0.0}
    for section, items in REPORT_STRUCTURE:
        for item in items:
            current = grouped.get((year, month, section, item), {"qty": 0.0, "amount": 0.0})
            cumulative = {"qty": 0.0, "amount": 0.0}
            for mm in range(1, month + 1):
                values = grouped.get((year, mm, section, item), {"qty": 0.0, "amount": 0.0})
                cumulative["qty"] += values["qty"]
                cumulative["amount"] += values["amount"]
            current_rows[(section, item)] = current
            cumulative_rows[(section, item)] = cumulative
            total_current["qty"] += current["qty"]
            total_current["amount"] += current["amount"]
            total_cumulative["qty"] += cumulative["qty"]
            total_cumulative["amount"] += cumulative["amount"]
    return current_rows, cumulative_rows, total_current, total_cumulative


def write_month_sheet(ws, month: int, current_year: int, previous_year: int, grouped: dict[tuple[int, int, str, str], dict[str, float]]) -> None:
    ws.append(["1. 제품별 판매현황"])
    ws.append(["구분", "", f"{month}월 당월", "", "", "", "", "", "", "", f"{month}월 누계", "", "", "", "", "", "", ""])
    ws.append(["", "", "수량", "전년", "증감", "구성비", "금액", "전년", "YOY", "구성비", "수량", "전년", "YOY", "구성비", "금액", "전년", "YOY", "구성비"])
    style_header(ws, 2)
    style_header(ws, 3)

    current_rows, cumulative_rows, total_current, total_cumulative = extract_report_rows(grouped, current_year, month)
    prev_rows, prev_cum_rows, total_prev, total_prev_cum = extract_report_rows(grouped, previous_year, month)

    for section, items in REPORT_STRUCTURE:
        first = True
        sec_cur_qty = sec_prev_qty = sec_cur_amt = sec_prev_amt = 0.0
        sec_cum_qty = sec_prev_cum_qty = sec_cum_amt = sec_prev_cum_amt = 0.0
        for item in items:
            cur = current_rows[(section, item)]
            prev = prev_rows[(section, item)]
            cur_cum = cumulative_rows[(section, item)]
            prev_cum = prev_cum_rows[(section, item)]
            sec_cur_qty += cur["qty"]; sec_prev_qty += prev["qty"]; sec_cur_amt += cur["amount"]; sec_prev_amt += prev["amount"]
            sec_cum_qty += cur_cum["qty"]; sec_prev_cum_qty += prev_cum["qty"]; sec_cum_amt += cur_cum["amount"]; sec_prev_cum_amt += prev_cum["amount"]
            ws.append([section if first else None, item, cur["qty"] or None, prev["qty"] or None, safe_ratio(cur["qty"], prev["qty"]), safe_share(cur["qty"], total_current["qty"]), cur["amount"] or None, prev["amount"] or None, safe_ratio(cur["amount"], prev["amount"]), safe_share(cur["amount"], total_current["amount"]), cur_cum["qty"] or None, prev_cum["qty"] or None, safe_ratio(cur_cum["qty"], prev_cum["qty"]), safe_share(cur_cum["qty"], total_cumulative["qty"]), cur_cum["amount"] or None, prev_cum["amount"] or None, safe_ratio(cur_cum["amount"], prev_cum["amount"]), safe_share(cur_cum["amount"], total_cumulative["amount"])])
            first = False
        ws.append([None, "소계", sec_cur_qty or None, sec_prev_qty or None, safe_ratio(sec_cur_qty, sec_prev_qty), safe_share(sec_cur_qty, total_current["qty"]), sec_cur_amt or None, sec_prev_amt or None, safe_ratio(sec_cur_amt, sec_prev_amt), safe_share(sec_cur_amt, total_current["amount"]), sec_cum_qty or None, sec_prev_cum_qty or None, safe_ratio(sec_cum_qty, sec_prev_cum_qty), safe_share(sec_cum_qty, total_cumulative["qty"]), sec_cum_amt or None, sec_prev_cum_amt or None, safe_ratio(sec_cum_amt, sec_prev_cum_amt), safe_share(sec_cum_amt, total_cumulative["amount"])])

    ws.append(["합계", None, total_current["qty"] or None, total_prev["qty"] or None, safe_ratio(total_current["qty"], total_prev["qty"]), 1 if total_current["qty"] else None, total_current["amount"] or None, total_prev["amount"] or None, safe_ratio(total_current["amount"], total_prev["amount"]), 1 if total_current["amount"] else None, total_cumulative["qty"] or None, total_prev_cum["qty"] or None, safe_ratio(total_cumulative["qty"], total_prev_cum["qty"]), 1 if total_cumulative["qty"] else None, total_cumulative["amount"] or None, total_prev_cum["amount"] or None, safe_ratio(total_cumulative["amount"], total_prev_cum["amount"]), 1 if total_cumulative["amount"] else None])
    ws.append([None, "이카운트 총 금액", total_current["qty"] or None, total_prev["qty"] or None, None, None, total_current["amount"] or None, total_prev["amount"] or None, None, None, total_cumulative["qty"] or None, total_prev_cum["qty"] or None, None, None, total_cumulative["amount"] or None, total_prev_cum["amount"] or None, None, None])


def build_workbook() -> None:
    master_rows = load_master_rows(BASE_PATH)
    rule_rows = load_rule_rows(BASE_PATH)
    master_lookup = build_master_lookup(master_rows)
    raw_records = parse_product_records(RAW_PRODUCT_PATH)
    normalized_records: list[NormalizedRecord] = []
    for record in raw_records:
        normalized_records.extend(expand_record(record, master_lookup, rule_rows))
    validation_records = parse_validation_records(VALIDATION_PATH)
    if not validation_records:
        validation_records = aggregate_validation(normalized_records)
    mapping_rows = build_mapping_rows(normalized_records)
    report_grouped = build_report_aggregates(normalized_records)

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "사용안내"
    info_rows = [
        ("시트", "설명"),
        ("기준_상품마스터", "사용자가 입력한 소비자가 기준표 복사본"),
        ("기준_세트규칙", "세트 분해 기준"),
        ("기준_부서", "소비자가 기준 부서 / 원본금액 부서 구분"),
        ("기준_보고항목매핑", "품목코드별 최종 보고서 행 매핑"),
        ("원본_품목월별", "품목 x 부서 x 월 원천 집계"),
        ("검증_부서월별", "부서별 월 합계 검증표"),
        ("계산_정규화", "소비자가 기준 3개 부서 정산수량/비교매출 계산 결과"),
        ("결과_월별보고", "월별 보고서 생성 요약"),
    ]
    for row in info_rows:
        ws_info.append(row)
    if not VALIDATION_PATH.exists():
        ws_info.append(("검증 참고", "qiDJlwun3rUVrKwG.xlsx 파일을 찾지 못해 원천 데이터 집계로 검증 시트를 대체함"))
    style_header(ws_info)

    src_wb = openpyxl.load_workbook(BASE_PATH, read_only=True, data_only=True)
    ws_master = wb.create_sheet("기준_상품마스터")
    copy_sheet_values(src_wb["상품마스터"], ws_master)
    style_header(ws_master)

    ws_rules = wb.create_sheet("기준_세트규칙")
    copy_sheet_values(src_wb["세트규칙"], ws_rules)
    style_header(ws_rules)

    ws_dept = wb.create_sheet("기준_부서")
    dept_rows = [
        ["부서명", "비교 기준", "세트 분해", "매출 계산 방식", "비고"],
        ["트로이아르케사업팀", "소비자가", "Y", "정산수량 x 소비자가", "무상출고 제외"],
        ["셀럽B2B", "소비자가", "Y", "정산수량 x 소비자가", "무상출고 제외"],
        ["밀리맘사업팀", "소비자가", "Y", "정산수량 x 소비자가", "무상출고 제외"],
        ["기타부서", "원본금액", "N", "원본 수량/금액 그대로 사용", "무상출고 제외"],
    ]
    for row in dept_rows:
        ws_dept.append(row)
    style_header(ws_dept)

    ws_map = wb.create_sheet("기준_보고항목매핑")
    ws_map.append(["품목코드", "품목명", "라인", "보고대분류", "보고항목", "기타분류여부", "비고"])
    for row in sorted(mapping_rows, key=lambda item: (item[3], item[4], item[0])):
        ws_map.append(row)
    style_header(ws_map)

    ws_raw = wb.create_sheet("원본_품목월별")
    ws_raw.append(["기준월", "연도", "월", "거래처그룹2코드", "거래처그룹2", "품목코드", "품목명", "수량", "공급가액", "부가세", "합계"])
    for record in raw_records:
        ws_raw.append([record.month_key, record.year, record.month, record.dept_code, record.dept_name, record.code, record.name, record.qty, record.supply, record.vat, record.total])
    style_header(ws_raw)

    ws_validation = wb.create_sheet("검증_부서월별")
    ws_validation.append(["기준월", "부서", "수량", "비교금액", "출처"])
    source_label = VALIDATION_PATH.name if VALIDATION_PATH.exists() else "원천데이터 집계"
    for row in validation_records:
        ws_validation.append([row["month_key"], row["dept_name"], row["qty"], row["total"], source_label])
    style_header(ws_validation)

    ws_norm = wb.create_sheet("계산_정규화")
    ws_norm.append(["기준월", "연도", "월", "부서", "원본품목코드", "원본품목명", "정규화품목코드", "정규화품목명", "라인", "구분", "보고대분류", "보고항목", "원본수량", "실물수량", "정산수량", "소비자가", "비교매출", "원본합계", "계산방식", "기타여부", "비고"])
    for record in normalized_records:
        ws_norm.append([record.month_key, record.year, record.month, record.dept_name, record.source_code, record.source_name, record.target_code, record.target_name, record.line, record.division, record.report_group, record.report_item, record.source_qty, record.physical_qty, record.settle_qty, record.retail_price, record.compare_amount, record.raw_total, record.calc_method, record.is_other, record.note])
    style_header(ws_norm)

    ws_result = wb.create_sheet("결과_월별보고")
    ws_result.append(["월", "설명"])
    months = sorted({record.month for record in normalized_records if record.year == 2026})
    for month in months:
        ws_result.append([f"{month}월", f"{month}월 시트를 확인하세요"])
    style_header(ws_result)

    for month in months:
        write_month_sheet(wb.create_sheet(f"{month}월"), month, 2026, 2025, report_grouped)

    for ws in wb.worksheets:
        auto_width(ws)
        ws.freeze_panes = "A2"

    for row in ws_map.iter_rows(min_row=2):
        if row[5].value == "Y":
            for cell in row:
                cell.fill = WARN_FILL

    for row in ws_rules.iter_rows(min_row=2):
        if row[9].value == "검토필요":
            for cell in row:
                cell.fill = WARN_FILL
        else:
            for cell in row:
                cell.fill = GOOD_FILL

    wb.save(OUTPUT_PATH)
    print(f"완료: {OUTPUT_PATH}")
    print(f"원본 레코드: {len(raw_records)}")
    print(f"정규화 레코드: {len(normalized_records)}")
    print(f"보고항목 매핑 수: {len(mapping_rows)}")


if __name__ == "__main__":
    build_workbook()
